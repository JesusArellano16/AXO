import shutil
import os
import datetime as dt
import openpyxl
from copy import copy
from collections import Counter
from time import sleep
import openpyxl.styles
from openpyxl.utils import get_column_letter
from new_queries import new_queries
import time
import shutil
from openpyxl.utils import quote_sheetname


name = 'Reporte_Discovery'
origin_path = f'./src/{name}.xlsx'
#central = 'IXTLA'
current_date_and_time = str(dt.date.today())

def copy_style_from_row(ws, source_row, start_row, end_row):
    for row in range(start_row, end_row + 1):
        for col in range(1, ws.max_column + 1):
            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=row, column=col)

            # Copiar estilos individuales
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)


def get_strictly_unique_combinations(central):
    path = r'./ARCHIVOS_REPORTES/'+central+ r'/'+current_date_and_time
    des_path = path + r'/' + f'EOL_{central}_{current_date_and_time}.xlsx'
    wb = openpyxl.load_workbook(des_path, data_only=True)
    ws = wb.active

    combos = []

    # Iterate over rows and extract values from B, G, H, I (columns 2, 7, 8, 9)
    for row in ws.iter_rows(min_row=2, max_col=11):  # Read up to column I
        b = row[1].value  # Column B (index 1)
        g = row[6].value  # Column G (index 6)
        h = row[7].value  # Column H (index 7)
        i = row[8].value  # Column I (index 8)
        j = row[9].value  # Column I (index 9)
        k = row[10].value  # Column I (index 10)
        combo = (b, g, h, i,j,k)
        combos.append(combo)

    # Contar cuántas veces aparece cada combinación
    counter = Counter(combos)

    return counter


def eol(central):
    #counter, strictly_unique_combos = get_strictly_unique_combinations(central=central)
    counter = get_strictly_unique_combinations(central=central)
    path_rep = r'./ARCHIVOS_REPORTES/'+central+ r'/'+current_date_and_time
    des_path_rep = path_rep + r'/' + f'Reporte_Discovery_{central}_{current_date_and_time}.xlsx'
    wb_rep = openpyxl.load_workbook(des_path_rep)
    ws_rep = wb_rep[f'Inventario - EOL']
    aux = 1
    for combo, count in counter.items():
        ws_rep[f'A{aux+5}'].value = aux
        ws_rep[f'B{aux+5}'].value = combo[0]
        ws_rep[f'C{aux+5}'].value = combo[1]
        ws_rep[f'D{aux+5}'].value = combo[2]
        ws_rep[f'E{aux+5}'].value = combo[3]
        ws_rep[f'F{aux+5}'].value = combo[4]
        ws_rep[f'G{aux+5}'].value = combo[5]
        ws_rep[f'H{aux+5}'].value = count
        aux += 1
    copy_style_from_row(ws=ws_rep,source_row=6,start_row=7,end_row=len(counter)+5)

    temp_path = des_path_rep.replace(".xlsx", "_temp.xlsx")
    wb_rep.save(temp_path)
    wb_rep.close()
    if os.path.exists(des_path_rep):
        os.remove(des_path_rep)  # Borrar el original (si existe)
    shutil.move(temp_path, des_path_rep)  # Renombrar el temporal al original
    return len(counter)



def totalAssets(central,file,serv):
    path = r'./ARCHIVOS_REPORTES/'+central+ r'/'+current_date_and_time
    des_path = path + r'/' + f'{file}_{central}_{current_date_and_time}.xlsx'
    wb = openpyxl.load_workbook(des_path, data_only=True)
    ws = wb.active
    aux = 0
    cortex = 0
    vp = 0
    if serv:
        for row in range(1, ws.max_row + 1):
            valor_f = ws[f'F{row}'].value
            valor_g = ws[f'G{row}'].value

            if valor_f == "SI":
                cortex += 1
            elif valor_f == "NO" and valor_g == "SI":
                vp += 1
        
    for cell in ws['A']:
        if cell.value not in (None,""):
            aux += 1
    wb.close()
    if aux == 0:    return 0,cortex,vp
    else:   return aux-1,cortex,vp

def getFormat(ws_rep, col, row,beg):
        prev_cell = ws_rep[f'{col}{row+beg}'] 
        new_cell = ws_rep[f'{col}{row+beg+1}']

        new_cell.font = copy(prev_cell.font)
        new_cell.fill = copy(prev_cell.fill)
        new_cell.border = copy(prev_cell.border)
        new_cell.alignment = copy(prev_cell.alignment)
        new_cell.number_format = copy(prev_cell.number_format)

def pcs_Inv(central,file, sheet):   
    unsupported_oses = [
    "Oracle Solaris","Windows Server 2008 R2",
    "Windows Server 2003 R2","SunOS 10","SunOS 11.1",
    "SunOS 11.2","SunOS 11.3","IBM AIX 6.1","IBM AIX 7.1",
    "IBM AIX 7.2","IBM AIX 5.3","SunOS 9","SunOS 11.4.23.69.3",
    "SunOS 11.0","SunOS 11.4","SunOS 11.4.0.15.0"
    ] 
    path_rep = r'./ARCHIVOS_REPORTES/'+central+ r'/'+current_date_and_time
    des_path_rep = path_rep + r'/' + f'Reporte_Discovery_{central}_{current_date_and_time}.xlsx'
    wb_rep = openpyxl.load_workbook(des_path_rep)
    ws_rep = wb_rep[f'{sheet}']
    
    path = r'./ARCHIVOS_REPORTES/'+central+ r'/'+current_date_and_time
    des_path = path + r'/' + f'{file}_{central}_{current_date_and_time}.xlsx'
    wb = openpyxl.load_workbook(des_path, data_only=True)
    ws = wb.active
    aux = False
    for row in range(2, ws.max_row + 1):
        if aux:
            for col in ['A', 'B', 'C', 'D', 'E', 'F','G', 'H', 'I', 'J']:
                getFormat(ws_rep, col, row,beg=3)
        ws_rep[f'A{row+4}'].value = row-1
        ws_rep[f'B{row+4}'].value = ws[f'B{row}'].value
        ws_rep[f'C{row+4}'].value = ws[f'C{row}'].value
        ws_rep[f'D{row+4}'].value = ws[f'D{row}'].value
        ws_rep[f'E{row+4}'].value = ws[f'E{row}'].value
        ws_rep[f'F{row+4}'].value = ws[f'F{row}'].value
        ws_rep[f'G{row+4}'].value = ws[f'G{row}'].value
        conditions = ', '.join([f'E{row+4}="{os}"' for os in unsupported_oses])
        #formula = f'=IF(OR({conditions}), "Cortex not supported", "NA")'
        formula = (
            f'=IF(F{row+4}="SI", "NA", '
            f'IF(OR({conditions}), "Cortex not supported", "NA"))'
        )
        if file == 'SERVERS':
            ws_rep[f'J{row+4}'].value = formula

        aux = True
    temp_path = des_path_rep.replace(".xlsx", "_temp.xlsx")
    wb_rep.save(temp_path)
    wb_rep.close()
    wb.close()
    if os.path.exists(des_path_rep):
        os.remove(des_path_rep)  # Borrar el original (si existe)
    shutil.move(temp_path, des_path_rep)  # Renombrar el temporal al original


def get_strictly_unique_combinations_vuln(central,vuln):
    path = r'./ARCHIVOS_REPORTES/'+central+ r'/'+current_date_and_time
    des_path = path + r'/' + f'{vuln}_SEV_{central}_{current_date_and_time}.xlsx'
    wb = openpyxl.load_workbook(des_path, data_only=True)
    ws = wb[f'{vuln}_SEV_{central}_{current_date_and_time}']

    combos = []

    # Iterate over rows and extract values from B, G, H, I (columns 2, 7, 8, 9)
    for row in ws.iter_rows(min_row=2, max_col=7):  # Read up to column I
        b = row[5].value  # Column B (index 1)
        combo = (b)
        combos.append(combo)

    # Contar cuántas veces aparece cada combinación
    counter = Counter(combos)
    return counter

def vuln(vulnerabilitie,central,path_rep):
    path_rep = r'./ARCHIVOS_REPORTES/'+central+ r'/'+current_date_and_time
    des_path = path_rep + r'/' + f'{vulnerabilitie}_SEV_{central}_{current_date_and_time}.xlsx'
    if not os.path.exists(des_path):
        path_rep = path_rep + r'/' + f'Reporte_Discovery_{central}_{current_date_and_time}.xlsx'
        wb = openpyxl.load_workbook(path_rep)
        ws = wb[f'Inventario']
        for row in range(6, ws.max_row+1):
            ws[f'H{row}'].value = 0
            ws[f'I{row}'].value = 0
        if vulnerabilitie == 'CRITICAL':
            sheet_reporte = wb[f'Resumen']
            sheet_reporte['E20'].value = 0
        elif vulnerabilitie == 'HIGH':
            sheet_reporte = wb[f'Resumen']
            sheet_reporte['E21'].value = 0

        wb.save(path_rep)
        wb.close()
        return
    path_reporte = path_rep + r'/' + f'Reporte_Discovery_{central}_{current_date_and_time}.xlsx'
    file_reporte = openpyxl.load_workbook(path_reporte)
    sheet_reporte = file_reporte[f'Inventario']

    counter = get_strictly_unique_combinations_vuln(central=central, vuln=vulnerabilitie)
    
    for row_rep in range(6, sheet_reporte.max_row+1):
        aux = False
        for combo, count in counter.items():
            if sheet_reporte[f'B{row_rep}'].value == combo:
                if vulnerabilitie == 'CRITICAL':
                    sheet_reporte[f'H{row_rep}'].value = count
                    aux = True
                elif vulnerabilitie == 'HIGH':
                    sheet_reporte[f'I{row_rep}'].value = count
                    aux = True
        if not aux:
            if vulnerabilitie == 'CRITICAL':
                sheet_reporte[f'H{row_rep}'].value = 0
            elif vulnerabilitie == 'HIGH':
                sheet_reporte[f'I{row_rep}'].value = 0

    
    file_reporte.save(path_reporte)
    file_reporte.close()

def esperar_archivos(central, fecha, carpeta_base="./ARCHIVOS_REPORTES"):
    flag_path = [
        f'./ARCHIVOS_REPORTES/{central}/{fecha}/done/critical_{central}.done',
        f'./ARCHIVOS_REPORTES/{central}/{fecha}/done/high_{central}.done',
        f'./ARCHIVOS_REPORTES/{central}/{fecha}/done/eol_{central}.done',
        f'./ARCHIVOS_REPORTES/{central}/{fecha}/done/NET_DEV_{central}.done',
        f'./ARCHIVOS_REPORTES/{central}/{fecha}/done/SERVERS_{central}.done',
        f'./ARCHIVOS_REPORTES/{central}/{fecha}/done/TOTAL_ASSETS_{central}.done',
        f'./ARCHIVOS_REPORTES/{central}/{fecha}/done/PCs_{central}.done',
    ]
    while True:
        if all(os.path.exists(path) for path in flag_path):
            break
        time.sleep(1)  # Esperar 1 segundo antes de volver a revisar

def verificar_archivo(central):
    # Ruta de la carpeta donde debe estar
    ruta_base = "AXONIUS_FILES"
    ruta_completa = os.path.join(ruta_base, central)

    # Verificar si la carpeta central existe
    if not os.path.isdir(ruta_completa):
        return False  # La carpeta no existe

    # Verificar si el archivo "vuln.csv" está dentro de central
    return f"eol_{central}.csv" in os.listdir(ruta_completa)

def Report(central2):
    central = central2.nombre
    full_Name = central2.fullName
    from table import mostrar_tabla
    from centrales  import centrales
    mostrar_tabla(centrales, current_date_and_time)
    esperar_archivos(
        central=central,
        fecha=current_date_and_time,
    )
    path = r'./ARCHIVOS_REPORTES/'+central
    if not os.path.exists(path):
        os.mkdir(path)
    path = path + r'/'+current_date_and_time
    if not os.path.exists(path):
        os.mkdir(path)
    des_path = path + r'/' + f'{name}_{central}_{current_date_and_time}.xlsx'
    shutil.copy(origin_path, des_path)

    wb = openpyxl.load_workbook(des_path)
    ws = wb['Inventario - Red']
    path_aux = path + r'/' + f'NET_DEV_{central}_{current_date_and_time}.xlsx'
    wb_aux = openpyxl.load_workbook(path_aux)
    ws_aux = wb_aux.active

    # Limpia de la fila 2 hacia abajo
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    # Recorrer filas y columnas
    for row in ws_aux.iter_rows(min_row=1):
        for cell in row:
            # Ajustar fila de destino para que empiece también desde la fila 2
            target_row = cell.row + 3 # ya empieza desde 2, no hay que ajustar nada más
            ws.cell(row=target_row, column=cell.column, value=cell.value)

    # Determinar la última fila pegada
    last_row_written = ws_aux.max_row + 3  # porque comenzamos desde fila 1 y pegamos desde fila 2

    # Aplicar estilo de la fila 3 a todas las filas copiadas
    copy_style_from_row(ws, source_row=5, start_row=4, end_row=last_row_written)
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f"{col}4"].font = openpyxl.styles.Font(bold=True)
    fil = "A4:F4"
    ws.auto_filter.ref = fil
    
    ws['A3'].value = f"Network Devices {full_Name} - Inventario "

    # Ajustar anchos de columna (opcional)
    for col in ws_aux.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = ws_aux.column_dimensions[col_letter].width
    wb.save(des_path)

    ws = wb[f'Resumen']
    ws['A2'].value = f'Inventario {full_Name} - Resumen'
    ws = wb[f'Inventario - PC']
    ws['A3'].value = f'PCs {full_Name} - Inventario'
    ws = wb[f'Inventario']
    ws['A3'].value = f'Servidores {full_Name} - Inventario'
    ws = wb[f'Inventario - EOL']
    ws['A3'].value = f'Servidores EOL {full_Name} - Inventario'


    t_assets,cortex,vp = totalAssets(central=central,file='TOTAL_ASSETS',serv=False)
    ws = wb[f'Resumen']
    ws['F3'].value = t_assets
    t_serv,cortex,vp = totalAssets(central=central,file='SERVERS', serv=True)
    ws['F5'].value = t_serv
    ws['F6'].value = cortex
    #ws['F7'].value = vp #VIRTUAL PATCHING
    #ws['F8'].value = t_serv - cortex - vp #NO PROTEGIDOS
    ws['F7'].value = t_serv - cortex
    t_pcs,cortex,vp = totalAssets(central=central,file='PCs', serv=True)
    ws['F10'].value = t_pcs
    ws['F11'].value = cortex
    ws['F12'].value = t_pcs - cortex
    t_net,cortex,vp = totalAssets(central=central,file='NET_DEV', serv=False)
    ws['F13'].value = t_net
    wb.save(des_path)
    wb.close()
    pcs_Inv(central=central,file='PCs',sheet='Inventario - PC')
    pcs_Inv(central=central,file='SERVERS',sheet='Inventario')

    if not verificar_archivo(central=central):
        wb = openpyxl.load_workbook(des_path)
        ws = wb[f'Resumen']
        ws['E22'].value = 0
        ws = wb[f'Inventario -EOL']
        ws['A6'].value = ''
        ws['B6'].value = ''
        ws['C6'].value = ''
        ws['D6'].value = ''
        ws['E6'].value = ''
        ws['F6'].value = ''
        ws['G6'].value = ''
        ws['H6'].value = ''
        wb.save(des_path)
        wb.close()
    else:
        eol_total = eol(central=central)

        wb = openpyxl.load_workbook(des_path)
        ws = wb[f'Resumen']
        ws['E22'].value = eol_total
        wb.save(des_path)
        wb.close()



    vuln(vulnerabilitie = 'CRITICAL', central = central, path_rep = des_path)
    vuln(vulnerabilitie = 'HIGH', central = central, path_rep = des_path)
    path_rep = r'./ARCHIVOS_REPORTES/'+central+ r'/'+current_date_and_time
    path_rep = path_rep + r'/' + f'Reporte_Discovery_{central}_{current_date_and_time}.xlsx'
    wb = openpyxl.load_workbook(path_rep)
    sheet_reporte = wb[f'Resumen']

    sheet_reporte['A19'].value = f'Servidores {full_Name} - Vulnerabilidades'

    sheet_reporte['E20'].value = f"=COUNTIF('Inventario'!H6:H1048576,\">0\")"
    sheet_reporte['E21'].value = f"=COUNTIF('Inventario'!I6:I1048576,\">0\")"
    wb.save(path_rep)
    wb.close()
    new_queries(central,full_Name)
    done_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/done/Reporte_{central}.done'
    with open(done_path, 'w') as f:
        f.write("done")

    if central == 'IXTLA':
        pass
    else:
        try:
            wb = openpyxl.load_workbook(path_rep)
            del wb['Adaptadores integrados']
            ## Se agrega guardado de eliminación de hoja de responsables
            wb.save(path_rep)
            wb.close()
        except:
            pass


    from table import mostrar_tabla
    from centrales  import centrales
    mostrar_tabla(centrales, current_date_and_time)