#!/usr/bin/env python3

import sys
from pathlib import Path
from datetime import datetime
from openpyxl.utils import get_column_letter
import xlwings as xw
from copy import copy
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import csv
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference

AGREGAR_SITIO = True

# Centrales recibidas como argumentos
centrales = sys.argv[1:]

if not centrales:
    print("No se recibieron centrales.")
    sys.exit(1)

# Fecha actual
hoy = datetime.now()
fecha_str = hoy.strftime("%Y-%m-%d")
anio = hoy.strftime("%Y")

meses_es = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre"
}

mes = meses_es[hoy.month]

# Ruta del directorio DESARROLLO
base_path = Path(__file__).resolve().parent.parent

# Ruta donde deben existir los reportes
reportes_path = (
    base_path
    / "REPORTES_SEMANALES"
    / anio
    / mes
    / fecha_str
)

faltantes = []
def sort_key(c):
    c = c.upper()

    # prioridad 1: R1-R9
    if c.startswith("R"):
        try:
            num = int(c[1:])
            return (0, num, "")
        except:
            pass

    # resto alfabético
    return (1, c)
    
centrales = sorted(centrales, key=sort_key)
for central in centrales:
    archivo = reportes_path / f"Reporte_Discovery_{central}_{fecha_str}.xlsx"

    if not archivo.exists():
        faltantes.append(archivo.name)

if faltantes:
    print("\nFALTA REPORTE\n")
    print("Reportes faltantes:")
    for reporte in faltantes:
        print(f"  - {reporte}")
    sys.exit(1)

from shutil import copy2
from openpyxl import load_workbook

# Tomar el reporte de R1 como plantilla
archivo_r1 = reportes_path / f"Reporte_Discovery_IXTLA_{fecha_str}.xlsx"

if not archivo_r1.exists():
    print(f"No existe {archivo_r1.name}")
    sys.exit(1)

# Nombre del nuevo reporte
archivo_destino = reportes_path / f"Reporte_Disc_Con_Clasificacion_{fecha_str}.xlsx"

# Duplicar archivo
copy2(archivo_r1, archivo_destino)

app = xw.App(visible=False)
wb_excel = app.books.open(str(archivo_destino))
ws_excel = wb_excel.sheets["Inventario"]

# Última fila con datos en columna A
last_row = ws_excel.range("A" + str(ws_excel.cells.last_cell.row)).end("up").row

# Convertir fórmulas a valores únicamente en J
rng = ws_excel.range(f"J6:J{last_row}")
valores = rng.value
rng.options(transpose=True).value = valores

wb_excel.save()
wb_excel.close()

# Volver a cargar el workbook para que openpyxl vea los cambios
wb = load_workbook(archivo_destino)


# ----------------------------
# Hoja Resumen
# ----------------------------
ws = wb["Resumen"]

ws["A2"] = "Inventario Con Clasificacion - Resumen"
ws["A19"] = "Servidores Con Clasificacion - Vulnerabilidades"

# ----------------------------
# Inventario - PC
# ----------------------------
ws = wb["Inventario - PC"]
ws["A3"] = "PCs Con Clasificacion - Inventario"

# ----------------------------
# Inventario - Red
# ----------------------------
ws = wb["Inventario - Red"]
ws["A3"] = "Network Devices Con Clasificacion - Inventario"

# ----------------------------
# Inventario
# ----------------------------
ws = wb["Inventario"]
ws["A3"] = "Servidores Con Clasificacion - Inventario"

# ----------------------------
# Inventario - EOL
# ----------------------------
ws = wb["Inventario - EOL"]
ws["A3"] = "Servidores EOL Con Clasificacion - Inventario"


# ----------------------------
# Eliminar hoja Data_R1
# ----------------------------
if "Data_IXTLA" in wb.sheetnames:
    ws_data = wb["Data_IXTLA"]
    wb.remove(ws_data)
    ws_data = wb["Adaptadores integrados"]
    wb.remove(ws_data)
else:
    print("No existe la hoja Data_R1")

# Libro destino (el de Adaptadores)
ws_dest = wb["Resumen"]

columna = 8   # J
app = xw.App(visible=False)
app.display_alerts = False

azul = PatternFill(
    fill_type="solid",
    start_color="0169C6",
    end_color="0169C6"
)

borde = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

font_header = Font(
    name="Calibri",
    size=18,
    bold=True,
    color="FFFFFF"
)

font_body_bold = Font(
    name="Calibri",
    bold=True
)

alineacion = Alignment(
    horizontal="center",
    vertical="center"
)
from openpyxl.utils import get_column_letter

def extender_merge_fila3(ws, col_sitio):

    # Guardar el valor de A3
    valor = ws["A3"].value

    # Buscar todos los merges que estén en fila 3 y deshacerlos
    merges_a_eliminar = []

    for rango in ws.merged_cells.ranges:
        if rango.min_row == 3 and rango.max_row == 3:
            merges_a_eliminar.append(str(rango))

    for rango in merges_a_eliminar:
        ws.unmerge_cells(rango)

    # Crear nuevo merge desde A3 hasta la columna SITIO
    letra_sitio = get_column_letter(col_sitio)

    ws.merge_cells(f"A3:{letra_sitio}3")

    # Restaurar texto
    ws["A3"] = valor
def append_hoja_fija(ws_dest, ws_src, sitio, fila_inicio=6):

    ultima_dest = ultima_fila_con_datos(ws_dest)
    ultima_src = ultima_fila_con_datos(ws_src)

    fila_destino = ultima_dest + 1

    # Número consecutivo
    siguiente_no = 1
    valor_ultimo_no = ws_dest.cell(ultima_dest, 1).value

    try:
        siguiente_no = int(valor_ultimo_no) + 1
    except:
        pass

    for fila_src in range(fila_inicio, ultima_src + 1):
        
        # ignorar filas vacías
        if all(ws_src.cell(fila_src, col).value is None for col in range(1, ws_src.max_column + 1)):
            continue

        copiar_formato_fila(ws_dest, fila_destino - 1, fila_destino)

        for col in range(1, ws_src.max_column + 1):
            if AGREGAR_SITIO:
                ws_dest.cell(
                    fila_destino,
                    COL_SITIO[ws_dest.title]
                ).value = sitio
            if col == 1:
                # Columna No
                ws_dest.cell(fila_destino, col).value = siguiente_no
            else:
                ws_dest.cell(fila_destino, col).value = \
                    ws_src.cell(fila_src, col).value

        siguiente_no += 1
        fila_destino += 1
def obtener_headers(ws, fila_header=4):
    headers = {}

    equivalencias = {
        "ADAPTERS": "ADAPTER",
        "COMMENTS": "COMMENT",
        "COMMENT": "COMMENT"
    }

    for col in range(1, ws.max_column + 1):

        valor = ws.cell(fila_header, col).value

        if valor:
            header = str(valor).strip().upper()

            if header in equivalencias:
                header = equivalencias[header]

            headers[header] = col

    return headers
def ultima_fila_con_datos(ws, col=1):
    fila = ws.max_row
    while fila > 1 and ws.cell(fila, col).value is None:
        fila -= 1
    return fila
def copiar_formato_fila(ws, fila_origen, fila_destino):
    for col in range(1, ws.max_column + 1):
        c1 = ws.cell(fila_origen, col)
        c2 = ws.cell(fila_destino, col)

        c2.font = copy(c1.font)
        c2.fill = copy(c1.fill)
        c2.border = copy(c1.border)
        c2.alignment = copy(c1.alignment)
        c2.number_format = copy(c1.number_format)
        c2.protection = copy(c1.protection)
COL_SITIO = {
    "Inventario - PC": 9,                  # I
    "Inventario - Red": 8,                 # H
    "Inventario": 11,                      # K
    "Inventario - EOL": 9,                 # I
    "Inventario No Identificados": 6,      # F
    "Inventario Identificados": 6          # F
}
def agregar_columna_sitio(ws, fila_header, fila_inicio, col_sitio, hacer_merge=False, sitio_inicial="IXTLA"):


    # Encabezado
    ws.cell(fila_header, col_sitio).value = "SITIO"
    # copiar formato de la columna izquierda
    if ws.title == "Inventario - PC":
        col_formato = 8   # H
    elif ws.title == "Inventario - Red":
        col_formato = 7   # G
    elif ws.title == "Inventario":
        col_formato = 10  # J
    elif ws.title == "Inventario - EOL":
        col_formato = 8   # H
    else:
        col_formato = col_sitio - 1
    if hacer_merge:

        # Formato de la fila superior (ej. H4)
        c_origen_4 = ws.cell(fila_header, col_formato)
        c_dest_4 = ws.cell(fila_header, col_sitio)

        c_dest_4.font = copy(c_origen_4.font)
        c_dest_4.fill = copy(c_origen_4.fill)
        c_dest_4.border = copy(c_origen_4.border)
        c_dest_4.alignment = copy(c_origen_4.alignment)
        c_dest_4.number_format = copy(c_origen_4.number_format)
        c_dest_4.protection = copy(c_origen_4.protection)

        # Formato de la fila inferior (ej. H5)
        c_origen_5 = ws.cell(fila_header + 1, col_formato)
        c_dest_5 = ws.cell(fila_header + 1, col_sitio)

        c_dest_5.font = copy(c_origen_5.font)
        c_dest_5.fill = copy(c_origen_5.fill)
        c_dest_5.border = copy(c_origen_5.border)
        c_dest_5.alignment = copy(c_origen_5.alignment)
        c_dest_5.number_format = copy(c_origen_5.number_format)
        c_dest_5.protection = copy(c_origen_5.protection)

        # Ahora sí hacer el merge
        ws.merge_cells(
            start_row=fila_header,
            start_column=col_sitio,
            end_row=fila_header + 1,
            end_column=col_sitio
        )

    

    c1 = ws.cell(fila_header, col_formato)
    c2 = ws.cell(fila_header, col_sitio)

    c2.font = copy(c1.font)
    c2.fill = copy(c1.fill)
    c2.border = copy(c1.border)
    c2.alignment = copy(c1.alignment)
    c2.number_format = copy(c1.number_format)
    c2.protection = copy(c1.protection)

    ultima = ultima_fila_con_datos(ws)

    ws.auto_filter.ref = (
        f"A{fila_header}:"
        f"{get_column_letter(col_sitio)}{ultima}"
    )

    for fila in range(fila_inicio, ultima + 1):

        c1 = ws.cell(fila, col_sitio - 1)
        c2 = ws.cell(fila, col_sitio)

        c2.font = copy(c1.font)
        c2.fill = copy(c1.fill)
        c2.border = copy(c1.border)
        c2.alignment = copy(c1.alignment)
        c2.number_format = copy(c1.number_format)
        c2.protection = copy(c1.protection)

        if ws.cell(fila, 1).value is not None:
            c2.value = sitio_inicial

    return col_sitio
def append_hoja(ws_dest, ws_src, fila_header, fila_inicio, sitio):
    headers_dest = obtener_headers(ws_dest, fila_header)
    headers_src = obtener_headers(ws_src, fila_header)
    siguiente_no = None

    if "NO" in headers_dest:
        col_no = headers_dest["NO"]

        ultima_dest = ultima_fila_con_datos(ws_dest)

        valor_actual = ws_dest.cell(ultima_dest, col_no).value

        try:
            siguiente_no = int(valor_actual) + 1
        except:
            siguiente_no = 1

    ultima_dest = ultima_fila_con_datos(ws_dest)
    fila_destino = ultima_dest + 1

    ultima_src = ultima_fila_con_datos(ws_src)
    #print(F'Copiando desde {ws_src} a {ws_dest}')

    for fila_src in range(fila_inicio, ultima_src + 1):

        if ws_src.cell(fila_src, 1).value is None:
            continue

        copiar_formato_fila(ws_dest, fila_destino - 1, fila_destino)

        if AGREGAR_SITIO:
            col_sitio = COL_SITIO[ws_dest.title]

        for header_dest, col_dest in headers_dest.items():
            if header_dest == "NO":
                if siguiente_no is not None:
                    ws_dest.cell(fila_destino, col_dest).value = siguiente_no
                continue

            if header_dest in headers_src:
                col_src = headers_src[header_dest]
                ws_dest.cell(fila_destino, col_dest).value = \
                    ws_src.cell(fila_src, col_src).value
            else:
                ws_dest.cell(fila_destino, col_dest).value = ""
        if AGREGAR_SITIO:
            ws_dest.cell(fila_destino, col_sitio).value = sitio
        fila_destino += 1

        if siguiente_no is not None:
            siguiente_no += 1   
    #print(F'Termino desde {ws_src} a {ws_dest}')

    # renumerar columna No
    if "NO" in headers_dest:
        col_no = headers_dest["NO"]

        numero = 1
        ultima = ultima_fila_con_datos(ws_dest)

        for fila in range(fila_inicio, ultima + 1):
            if ws_dest.cell(fila, col_no).value is not None:
                ws_dest.cell(fila, col_no).value = numero
                numero += 1
HOJAS = {
    "Inventario - PC": (4, 6),
    "Inventario - Red": (4, 5),
    "Inventario": (4, 6),
    "Inventario - EOL": (4, 6),
    "Inventario No Identificados": (4, 5),
    "Inventario Identificados": (4, 5)
}
MERGE_SITIO = {
    "Inventario - PC": True,
    "Inventario - Red": False,
    "Inventario": True,
    "Inventario - EOL": True,
    "Inventario No Identificados": False,
    "Inventario Identificados": False
}

if AGREGAR_SITIO:

    for nombre_hoja, (fila_header, fila_inicio) in HOJAS.items():

        if nombre_hoja in wb.sheetnames:

            ws_tmp = wb[nombre_hoja]

            col_sitio = agregar_columna_sitio(
                ws_tmp,
                fila_header,
                fila_inicio,
                COL_SITIO[nombre_hoja],
                MERGE_SITIO[nombre_hoja],
                "IXTLA"
            )

            extender_merge_fila3(ws_tmp, col_sitio)

for central in centrales:

    archivo_central = reportes_path / f"Reporte_Discovery_{central}_{fecha_str}.xlsx"

    # Abrir con data_only=True para obtener el resultado de las fórmulas
    wb_src = load_workbook(archivo_central, data_only=True)
    ws_src = wb_src["Resumen"]
    wb_excel = app.books.open(str(archivo_central))
    wb_excel.app.calculate()
    wb_excel.save()
    wb_excel.close()
    """
   with xw.App(visible=False) as app:
        wb_excel = app.books.open(str(archivo_central))

        # Fuerza el recálculo
        wb_excel.app.calculate()

        # Guarda el resultado de las fórmulas
        wb_excel.save()

        wb_excel.close()
        """
    if not archivo_central.exists():
        print(f"No existe {archivo_central.name}")
        continue



    letra_col = get_column_letter(columna)

    # Nombre de la central
    ws_dest.cell(row=2, column=columna).value = central
    celda = ws_dest.cell(row=2, column=columna)

    celda.fill = azul
    celda.font = font_header
    celda.border = borde
    celda.alignment = alineacion

    # G3:G16 <- F3:F16
    for fila in range(3, 17):
        dst = ws_dest.cell(row=fila, column=columna)
        src_style = ws_dest[f"F{fila}"]

        dst.value = ws_src[f"F{fila}"].value

        dst.font = copy(src_style.font)
        dst.fill = copy(src_style.fill)
        dst.border = copy(src_style.border)
        dst.alignment = copy(src_style.alignment)
        dst.number_format = copy(src_style.number_format)
        dst.protection = copy(src_style.protection)

    # G20:G22 <- E20:E22
    for fila in range(20, 23):
        dst = ws_dest.cell(row=fila, column=columna)

        dst.value = ws_src[f"E{fila}"].value

        dst.font = font_body_bold
        dst.border = borde
        dst.alignment = alineacion

    columna += 1
app.quit()

from openpyxl.utils import get_column_letter

col_inicio = 8  # J
col_fin = columna - 1  # última central

col_inicio_letter = "H"
col_fin_letter = "AI"
ws = wb["Resumen"]

for fila in range(3, 17):
    ws[f"F{fila}"] = f"=SUM({col_inicio_letter}{fila}:{col_fin_letter}{fila})"
for fila in range(20, 23):
    ws[f"E{fila}"] = f"=SUM({col_inicio_letter}{fila}:{col_fin_letter}{fila})"

# ----------------------------
# Formato encabezado Resumen
# ----------------------------
from copy import copy

ws = wb["Resumen"]

# Guardar estilo de A2 ANTES de deshacer el merge
estilo = {
    "font": copy(ws["A2"].font),
    "fill": copy(ws["A2"].fill),
    "border": copy(ws["A2"].border),
    "alignment": copy(ws["A2"].alignment),
    "number_format": ws["A2"].number_format,
    "protection": copy(ws["A2"].protection),
}

texto = ws["A2"].value

# Deshacer merges si existen
for rango in ("A1:F1", "A2:F2", "A1:F2"):
    try:
        ws.unmerge_cells(rango)
    except:
        pass

# Nuevo merge
ws.merge_cells("A1:F2")

# Restaurar contenido
ws["A1"] = texto

# Restaurar formato
ws["A1"].font = estilo["font"]
ws["A1"].fill = estilo["fill"]
ws["A1"].border = estilo["border"]
ws["A1"].alignment = estilo["alignment"]
ws["A1"].number_format = estilo["number_format"]
ws["A1"].protection = estilo["protection"]

for central in centrales:

    if central == "IXTLA":
        continue

    archivo_central = reportes_path / \
        f"Reporte_Discovery_{central}_{fecha_str}.xlsx"

    wb_src = load_workbook(archivo_central, data_only=True)

    for nombre_hoja, (fila_header, fila_inicio) in HOJAS.items():

        if nombre_hoja not in wb.sheetnames:
            continue

        if nombre_hoja not in wb_src.sheetnames:
            continue

        ws_dest = wb[nombre_hoja]
        ws_src = wb_src[nombre_hoja]

        if AGREGAR_SITIO:

            if nombre_hoja == "Inventario":
                append_hoja_fija(ws_dest, ws_src, central, fila_inicio=6)
            else:
                append_hoja(ws_dest, ws_src, fila_header, fila_inicio, central)

        else:

            if nombre_hoja == "Inventario":
                append_hoja_fija(ws_dest, ws_src, "", fila_inicio=6)
            else:
                append_hoja(ws_dest, ws_src, fila_header, fila_inicio, "")
    print(f'TEERMINAMOS DE COPIAR {central}')


from copy import copy
from openpyxl.utils import column_index_from_string

def mover_columna(ws, origen, destino):
    c_org = column_index_from_string(origen)
    c_dst = column_index_from_string(destino)

    for fila in range(1, ws.max_row + 1):
        src = ws.cell(fila, c_org)
        dst = ws.cell(fila, c_dst)

        dst.value = src.value
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)

        # Limpia la columna origen
        src.value = None

# Mover de derecha a izquierda
mover_columna(ws, "AA", "AD")
mover_columna(ws, "Z", "AA")
mover_columna(ws, "W", "AI")
mover_columna(ws, "V", "AH")
mover_columna(ws, "Q", "AG")
mover_columna(ws, "P", "AF")
mover_columna(ws, "O", "AC")
mover_columna(ws, "N", "Z")
mover_columna(ws, "M", "W")
mover_columna(ws, "X", "Q")
mover_columna(ws, "Y", "X")
mover_columna(ws, "T", "O")
mover_columna(ws, "L", "T")
mover_columna(ws, "R", "M")
mover_columna(ws, "S", "N")
mover_columna(ws, "Q", "R")
mover_columna(ws, "K", "Q")
mover_columna(ws, "I", "K")
mover_columna(ws, "M", "L")
mover_columna(ws, "O", "I")
mover_columna(ws, "N", "O")
mover_columna(ws, "J", "N")


from openpyxl.utils import column_index_from_string
from copy import copy

# Encabezados por región
regiones = [
    ("R1", "H"),
    ("R2", "K"),
    ("R3", "N"),
    ("R4", "Q"),
    ("R5", "T"),
    ("R6", "W"),
    ("R7", "Z"),
    ("R8", "AC"),
]

for nombre, col in regiones:

    c = column_index_from_string(col)

    # Merge de 3 columnas
    ws.merge_cells(
        start_row=1,
        start_column=c,
        end_row=1,
        end_column=c + 2
    )

    celda = ws.cell(1, c)
    celda.value = nombre

    celda.font = copy(estilo["font"])
    celda.fill = copy(estilo["fill"])
    celda.border = copy(estilo["border"])
    celda.alignment = copy(estilo["alignment"])
    celda.number_format = estilo["number_format"]
    celda.protection = copy(estilo["protection"])

# Encabezados individuales
individuales = {
    "AF": "R9",
    "AG": "CARSO",
    "AH": "IXTLA",
    "AI": "L_ALB",
}

for col, texto in individuales.items():

    celda = ws[col + "1"]
    celda.value = texto

    celda.font = copy(estilo["font"])
    celda.fill = copy(estilo["fill"])
    celda.border = copy(estilo["border"])
    celda.alignment = copy(estilo["alignment"])
    celda.number_format = estilo["number_format"]
    celda.protection = copy(estilo["protection"])
from copy import copy
# ----------------------------
# Fila 2: Cent / Ofic / CACs
# ----------------------------
bloques = ["H", "K", "N", "Q", "T", "W", "Z", "AC"]

for inicio in bloques:

    col = column_index_from_string(inicio)

    # Encabezados
    ws.cell(2, col).value = "Cent"
    ws.cell(2, col + 1).value = "Ofic"
    ws.cell(2, col + 2).value = "CACs"

    # Copiar formato de A2
    for offset in range(3):
        c = ws.cell(2, col + offset)

        c.font = copy(estilo["font"])
        c.fill = copy(estilo["fill"])
        c.border = copy(estilo["border"])
        c.alignment = copy(estilo["alignment"])
        c.number_format = estilo["number_format"]
        c.protection = copy(estilo["protection"])

    # Columna CACs -> "-" de fila 3 a 16
    for fila in range(3, 17):
        c = ws.cell(fila, col + 2)
        c.value = "-"
# R9, CARSO, IXTLA y L_ALB
ws["AF2"] = "Cent"
ws["AG2"] = "Ofic"
ws["AH2"] = "Site"
ws["AI2"] = "Site"

for col in ["AF", "AG", "AH", "AI"]:
    c = ws[f"{col}2"]

    c.font = copy(estilo["font"])
    c.fill = copy(estilo["fill"])
    c.border = copy(estilo["border"])
    c.alignment = copy(estilo["alignment"])
    c.number_format = estilo["number_format"]
    c.protection = copy(estilo["protection"])
from openpyxl.styles import Font

for fila in [8, 15]:
    for col in range(1, ws.max_column + 1):

        c = ws.cell(fila, col)

        c.font = Font(
            name=c.font.name,
            size=c.font.size,
            bold=c.font.bold,
            italic=c.font.italic,
            underline=c.font.underline,
            strike=c.font.strike,
            color="FF0000"
        )
from copy import copy
from openpyxl.utils import column_index_from_string

for destino in ["AB", "AE"]:
    for fila in range(3, 23):   # 3 a 22

        src = ws[f"AA{fila}"]
        dst = ws[f"{destino}{fila}"]

        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
# Eliminar merge si existe
try:
    ws.unmerge_cells("H19:AI19")
except:
    pass

# Crear merge
ws.merge_cells("H19:AI19")

c = ws["H19"]
c.value = "Servidores Con Clasificacion - Vulnerabilidades"

# Mismo formato que A1
c.font = copy(estilo["font"])
c.fill = copy(estilo["fill"])
c.border = copy(estilo["border"])
c.alignment = copy(estilo["alignment"])
c.number_format = estilo["number_format"]
c.protection = copy(estilo["protection"])
# Guardar cambios
wb.save(archivo_destino)


csv_path = base_path / "src" / "Graficas" / "data_ADAPTERS.csv"
ws_resumen = wb["Resumen"]

fecha_csv = hoy.strftime("%d/%m/%y")
def suma_fila(ws, fila, col_inicio=8, col_fin=None):
    total = 0

    for col in range(col_inicio, col_fin + 1):
        valor = ws.cell(fila, col).value

        if isinstance(valor, (int, float)):
            total += valor

    return total
col_fin = 35

nuevo_registro = {
    "Date": fecha_csv,
    "Total Dispositivos y/o IPs Descubiertos": suma_fila(ws_resumen, 3, 8, col_fin),
    "Dispositivos Identificados": suma_fila(ws_resumen, 4, 8, col_fin),
    "Servidores con Cortex": suma_fila(ws_resumen, 6, 8, col_fin),
    "Servidores no protegidos": suma_fila(ws_resumen, 7, 8, col_fin),
    "IPs No Identificados (Servidores)": suma_fila(ws_resumen, 15, 8, col_fin),
    "IPs No Identificados (No Vendor)": suma_fila(ws_resumen, 16, 8, col_fin),
    "Total Servidores": suma_fila(ws_resumen, 5, 8, col_fin)
}
if csv_path.exists():
    df = pd.read_csv(csv_path)

    # eliminar la fecha actual si existe
    df = df[df["Date"] != fecha_csv]

else:
    df = pd.DataFrame()

df = pd.concat([df, pd.DataFrame([nuevo_registro])], ignore_index=True)

# ordenar por fecha
df["Date_dt"] = pd.to_datetime(df["Date"], format="%d/%m/%y")
df = df.sort_values("Date_dt")
df = df.drop(columns=["Date_dt"])

df.to_csv(csv_path, index=False)

if "data_ADAPTADORES" in wb.sheetnames:
    del wb["data_ADAPTADORES"]

ws_data = wb.create_sheet("data_ADAPTADORES")
df = pd.read_csv(csv_path)

for fila in dataframe_to_rows(df, index=False, header=True):
    ws_data.append(fila)
chart = LineChart()

chart.title = "Tendencia Con Clasificacion"
chart.style = 2

chart.y_axis.title = "Cantidad"
chart.x_axis.title = "Fecha"

# unidades de 1000
chart.y_axis.majorUnit = 5000
data = Reference(
    ws_data,
    min_col=2,
    max_col=8,
    min_row=1,
    max_row=ws_data.max_row
)

cats = Reference(
    ws_data,
    min_col=1,
    min_row=2,
    max_row=ws_data.max_row
)

chart.add_data(data, titles_from_data=True)
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice
colors = [
    "FF0000",  # Total Dispositivos -> rojo
    "0000FF",  # Identificados -> azul
    "90EE90",  # Cortex -> verde claro
    "FFD700",  # No protegidos -> amarillo
    "800080",  # IPs servidores -> morado
    "00A86B",  # IPs no vendor -> verde fuerte
    "556B2F"   # Total servidores -> verde militar
]
for i, serie in enumerate(chart.series):
    if i < len(colors):
        serie.graphicalProperties.line.solidFill = colors[i]
        serie.graphicalProperties.line.width = 20000  # grosor opcional
chart.set_categories(cats)
chart.width = 40
chart.height = 18
ws_data.add_chart(chart, "J1")
ws_data.sheet_view.zoomScale = 110
ws_data.sheet_view.topLeftCell = "J1"
ws_data.sheet_view.selection[0].activeCell = "J1"
ws_data.sheet_view.selection[0].sqref = "J1"
wb.save(archivo_destino)