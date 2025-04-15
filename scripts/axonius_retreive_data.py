import axonius_api_client as axonapi  # Importar la librería de Axonius API Client
import json  # Importar la librería JSON para manejar datos en formato JSON
import warnings  # Importar warnings para manejar advertencias
import pandas as pd
import urllib3
import os
import numpy as np
import openpyxl
from openpyxl.styles import Alignment
from copy import copy
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.simplefilter("ignore",category=Warning)  # Ignorar advertencias para evitar mensajes innecesarios en la consola


def axonius_retreive_data(connect_args,saved_query_name,saved_query_name_clean,central,current_date_and_time):
    print(f'🚀 Working in {saved_query_name} ')
    path = r'./ARCHIVOS_REPORTES/'+central
    if not os.path.exists(path):
        os.mkdir(path)
    path = path + r'/'+current_date_and_time
    if not os.path.exists(path):
        os.mkdir(path)
    # Crear una instancia del cliente de Axonius
    client = axonapi.Connect(**connect_args)

    # Seleccionar el objeto API para dispositivos
    apiobj = client.devices

    # Obtener los dispositivos de la consulta guardada
    devices = apiobj.get_by_saved_query(saved_query_name)
    
    # Lista donde se almacenarán los dispositivos con datos filtrados
    clean_devices = []

    if saved_query_name_clean == "NET_DEV":
        clean_data = [
        "adapters",  # Adaptadores conectados al dispositivo
        "specific_data.data.hostname_preferred",  # Nombre de host preferido
        "specific_data.data.network_interfaces.ips_preferred",  # Dirección IP preferida
        "specific_data.data.network_interfaces.mac_preferred",  # Dirección MAC preferida
        "specific_data.data.os.type_distribution_preferred",  # Tipo de sistema operativo preferido
        "specific_data.data.network_interfaces.manufacturer"
        ]   
    # Definir los campos específicos que se desean extraer de cada dispositivo
    else:
        clean_data = [
            "adapters",  # Adaptadores conectados al dispositivo
            "specific_data.data.hostname_preferred",  # Nombre de host preferido
            "specific_data.data.network_interfaces.ips_preferred",  # Dirección IP preferida
            "specific_data.data.network_interfaces.mac_preferred",  # Dirección MAC preferida
            "specific_data.data.os.type_distribution_preferred",  # Tipo de sistema operativo preferido
        ]

    # Recorrer la lista de dispositivos obtenidos de la consulta guardada
    for device in devices:
        clean_devices_dict = {}  # Diccionario para almacenar datos filtrados del dispositivo
        for data in clean_data:
            try:
                clean_devices_dict[data] = device[data]  # Intentar obtener el valor del campo especificado
            except:
                pass  # Si el campo no existe en el dispositivo, ignorarlo
        if saved_query_name_clean == "NET_DEV":
            pass
        else:  
            if "paloalto_xdr_adapter" in clean_devices_dict["adapters"]:
                clean_devices_dict["CORTEX"] = "SI"
            if 'deep_security_adapter' in clean_devices_dict["adapters"]:
                clean_devices_dict["VIRTUAL PATCHING"] = "SI"
            if "paloalto_xdr_adapter" not in clean_devices_dict["adapters"]:
                clean_devices_dict["CORTEX"] = "NO"
            if 'deep_security_adapter' not in clean_devices_dict["adapters"]:
                clean_devices_dict["VIRTUAL PATCHING"] = "NO"
        clean_devices.append(clean_devices_dict)  # Agregar el diccionario con los datos filtrados a la lista

    # Guardar los datos filtrados en un archivo JSON
    with open(f"{saved_query_name_clean}.json", "w") as final:
        json.dump(clean_devices, final, indent=4)  # Guardar los datos con formato JSON legible
    
    with open(f"{saved_query_name_clean}.json", "r") as file:
        data = json.load(file)

    df = pd.DataFrame(data)
    df.to_excel(f'{saved_query_name}.xlsx',index=False, engine="openpyxl")

    df1 = pd.read_excel(f'{saved_query_name}.xlsx')
    if saved_query_name_clean == "NET_DEV":
        headers = {
            "adapters": "Adapter", 
            "specific_data.data.hostname_preferred": "Hostname", 
            "specific_data.data.network_interfaces.ips_preferred": "IPs", 
            "specific_data.data.network_interfaces.mac_preferred": "MAC",
            "specific_data.data.os.type_distribution_preferred": "OS",
            "specific_data.data.network_interfaces.manufacturer": "Manufacturer"
        }
    else:
        headers = {
            "adapters": "Adapter", 
            "specific_data.data.hostname_preferred": "Hostname", 
            "specific_data.data.network_interfaces.ips_preferred": "IPs", 
            "specific_data.data.network_interfaces.mac_preferred": "MAC",
            "specific_data.data.os.type_distribution_preferred": "OS"
        }
    #not_in_headers = False
    for key, value in headers.items():
        try:
            df1.rename(columns={key:value}, inplace=True)
        except:
            pass
    #df1[headers_old[0]] = df1[headers_old[0]].astype(str).str.replace(headers_old[0], headers_new[0])
    df1.to_excel(f'{saved_query_name}_modified.xlsx', index=False, engine="openpyxl")
    #print(not_in_headers)
    changes = {
        "[": "",
        "]": "",
        "'": "",
        ",": "\n"
    }

    df2 = pd.read_excel(f'{saved_query_name}_modified.xlsx')
    for key, value in headers.items():
        for k, v in changes.items():
            try:
                df2[value] = df2[value].astype(str).str.replace(k,v)
            except:
                pass
    df2.to_excel(f'{saved_query_name}_modified.xlsx', index=False, engine="openpyxl")



    archivo = f'{saved_query_name}_modified.xlsx'
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active  # Seleccionar la hoja activa
    if saved_query_name_clean == "NET_DEV" and central == "IXTLA":
        # Paso 1: Copiar los valores de la columna E
        col_e_values = [ws.cell(row=i, column=5).value for i in range(1, ws.max_row + 1)]

        # Paso 2: Insertar una nueva columna en la posición 2 (entre A y B)
        ws.insert_cols(2)

        # Paso 3: Pegar los valores de la columna E en esa nueva columna B
        for i, value in enumerate(col_e_values, start=1):
            ws.cell(row=i, column=2, value=value)

        ws["B1"].font = copy(ws["E1"].font)
        ws["B1"].fill = copy(ws["E1"].fill)
        ws["B1"].border = copy(ws["E1"].border)
        ws["B1"].alignment = copy(ws["E1"].alignment)
        ws["B1"].number_format = copy(ws["E1"].number_format)

        # Paso 4: Eliminar la columna E original (ahora desplazada a F)
        ws.delete_cols(6)

    columnas = [cell.value for cell in ws[1]]
    #print(columnas)
    if "MAC" in columnas:
        #print("TRUE")
        # Definir las columnas objetivo (A, C y D corresponden a índices 1, 3 y 4 en openpyxl)
        columnas_objetivo = [1, 3, 4]
        columnas_ancho = {
            "A": 30,
            "B": 50,
            "C": 20,
            "D": 20,
            "E": 20,
            "F": 10,
            "G": 17
        }
        fil = "A1:G1"
        if saved_query_name_clean == "NET_DEV" and central == "IXTLA":
            columnas_ancho["E"]=70
            fil = "A1:E1"
        if saved_query_name_clean == "NET_DEV" and central == "CARSO":
            columnas_ancho["F"]=70
            fil = "A1:F1"
        
    else:
        # Definir las columnas objetivo (A, C y D corresponden a índices 1, 3 y 4 en openpyxl)
        columnas_objetivo = [1, 3]
        columnas_ancho = {
            "A": 30,
            "B": 50,
            "C": 20,
            "D": 10,
            "E": 17,

        }
        fil = "A1:E1"
    


    # Aplicar ajuste de texto en las columnas A, C y D (desde la fila 2 en adelante)
    for col in columnas_objetivo:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)  # Activar "Wrap Text"

    for col, width in columnas_ancho.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = fil
    # Guardar el archivo modificado

    namew = saved_query_name_clean + '_' + central + "_" + str(current_date_and_time)
    name = './ARCHIVOS_REPORTES/'+central+'/'+current_date_and_time+'/' + namew  + r'.xlsx'
    wb.save(name)
    ss = openpyxl.load_workbook(name)
    ss_sheet = ss['Sheet1']
    ss_sheet.title = namew
    ss.save(name)
    remo = saved_query_name_clean + '.json'
    os.remove(f'{saved_query_name}_modified.xlsx')
    os.remove(f'{saved_query_name}.xlsx')
    os.remove(remo)
    wb = openpyxl.load_workbook(name)
    ws = wb[namew]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in ["nan", ""]:  
                cell.value = None  # Vaciar la celda
    wb.save(name)
    wb.close()  

    if 'PC' in saved_query_name_clean :
        noCortex = []
        wb = openpyxl.load_workbook(name)
        for fila in range(1, ws.max_row+1):
            celda = ws[f"F{fila}"]
            if celda.value and "NO" in str(celda.value):  # Verificar si el valor está en la celda
                deviceNo = []
                deviceNo.append(ws[f'B{fila}'].value)
                deviceNo.append(ws[f'C{fila}'].value)
                deviceNo.append(ws[f'D{fila}'].value)
                deviceNo.append("NO")
                deviceNo.append(fila)
                noCortex.append(deviceNo)
        for x in noCortex:
            x[2] = x[2].split("\n")
            #print(x[2])
            if len(x[2]) > 1: pass
            query = f'specific_data.data.network_interfaces.mac == "{x[2][0]}"'
            # Ejecutar la consulta
            assets = client.devices.get(query=query)
            ws = wb[namew]
            for asset in assets:
                if len(asset["specific_data.data.network_interfaces.ips"]) > 1 : pass
                if asset["specific_data.data.network_interfaces.ips"][0] == x[1] and asset["specific_data.data.network_interfaces.mac"][0] == x[2][0] and asset['specific_data.data.hostname'][0] != x[0]:
                    x[0] = x[0] + f"({asset['specific_data.data.hostname'][0]})"
                    ws[f'B{x[4]}'].value = x[0]
                    #print(ws[f'B{x[4]}'].value)
                    if "paloalto_xdr_adapter" in asset['adapters']:
                        x[3] = "SI"
                        ws[f'F{x[4]}'].value = "SI"
                    wb.save(name)
                elif (asset["specific_data.data.network_interfaces.ips"][0] == x[1] and asset["specific_data.data.network_interfaces.mac"][0] == x[2][0] and asset['specific_data.data.hostname'][0] != x[0]):
                    x[0] = x[0] + f"({asset['specific_data.data.hostname'][0]})"
                    ws[f'B{x[4]}'].value = x[0]
                    #print(ws[f'B{x[4]}'].value)
                    if "paloalto_xdr_adapter" in asset['adapters']:
                        x[3] = "SI"
                        ws[f'F{x[4]}'].value = "SI"
                    wb.save(name)
        wb.save(name)
        wb.close()
    print(f'✅{name} created')
