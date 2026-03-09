import axonius_api_client as axonapi
import os
import json
import urllib3
import datetime as dt
from dotenv import load_dotenv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from axonius_api_client.constants.fields import UnknownFieldSchema
import warnings

warnings.filterwarnings("ignore", category=UnknownFieldSchema)

# ======================================================
# ENV
# ======================================================
dotenv_path = Path(__file__).parent / ".env"
load_dotenv(dotenv_path=dotenv_path)

warnings.filterwarnings("ignore", category=axonapi.exceptions.ExtraAttributeWarning)
warnings.simplefilter("ignore", urllib3.exceptions.InsecureRequestWarning)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

connect_args = {
    "url": os.getenv("AXONIUS_URL"),
    "key": os.getenv("AXONIUS_KEY"),
    "secret": os.getenv("AXONIUS_SECRET"),
    "verify": False
}

# ======================================================
# PATH BASE
# ======================================================
base_path = Path(__file__).parent.parent

json_folder = base_path / "AXONIUS_FILES" / "GENERAL_JSON"
json_folder.mkdir(parents=True, exist_ok=True)

# Fecha
today = dt.date.today()
year = today.strftime("%Y")
meses = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre"
}

month_name = meses[today.month]
date_folder = today.strftime("%Y-%m-%d")
date_excel = today.strftime("%d%m%y")

excel_folder = base_path / "REPORTES_SEMANALES" / year / month_name / date_folder
excel_folder.mkdir(parents=True, exist_ok=True)

def retrieve_eol():
    client = axonapi.Connect(**connect_args)
    devices_api = client.devices

    query_name = "GENERAL ASSETS WITH EOL AND EOS"

    devices = devices_api.get_by_saved_query(query_name)

    # ======================================================
    # SAVE JSON
    # ======================================================
    json_output = json_folder / f"general_assets_eol_eos_{date_folder}.json"

    with open(json_output, "w", encoding="utf-8") as f:
        json.dump(devices, f, indent=2)


    # ======================================================
    # EXCEL FILE
    # ======================================================
    excel_file = excel_folder / f"EOL_&_EOS_SERVERS_{date_excel}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "SERVERS"

    headers = [
        "HOSTNAME",
        "IP",
        "MAC ADDRESS",
        "END-OF-LIFE",
        "END-OF-SUPPORT",
        "OS",
        "FABRICANTE",
        "SITIO",
        "CORTEX"
    ]

    # ======================================================
    # HEADER STYLE
    # ======================================================
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws.auto_filter.ref = "A1:I1"

    # ======================================================
    # COLUMN WIDTH
    # ======================================================
    widths = {
        "A": 40,
        "B": 25,
        "C": 30,
        "D": 30,
        "E": 30,
        "F": 40,
        "G": 40,
        "H": 15,
        "I": 15
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws["B1"].alignment = Alignment(wrap_text=True)
    ws["C1"].alignment = Alignment(wrap_text=True)

    # ======================================================
    # HELPERS
    # ======================================================
    def format_list(value, remove_parenthesis=False):
        if isinstance(value, list):
            value = "\n".join(str(v) for v in value)

        if isinstance(value, str) and remove_parenthesis:
            value = value.replace("(", "").replace(")", "")

        return value

    # ======================================================
    # INSERT DATA
    # ======================================================
    for d in devices:

        hostname = d.get("specific_data.data.hostname_preferred", "")

        ips = format_list(
            d.get("specific_data.data.network_interfaces.ips_v4_preferred", [])
        )

        mac = format_list(
            d.get("specific_data.data.network_interfaces.mac_preferred", [])
        )

        eol = d.get("specific_data.data.os.end_of_life_preferred", "")
        eos = d.get("specific_data.data.os.end_of_support_preferred", "")
        os_type = d.get("specific_data.data.os.type_distribution_preferred", "")

        fabricante = format_list(
            d.get("specific_data.data.network_interfaces.manufacturer", []),
            remove_parenthesis=True
        )

        sitio = format_list(d.get("labels", []))

        adapters = d.get("adapters", [])
        cortex = "SI" if "paloalto_xdr_adapter" in adapters else "NO"

        ws.append([
            hostname,
            ips,
            mac,
            eol,
            eos,
            os_type,
            fabricante,
            sitio,
            cortex
        ])

    # Wrap columnas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[1].alignment = Alignment(wrap_text=True)
        row[2].alignment = Alignment(wrap_text=True)

    # ======================================================
    # SAVE EXCEL
    # ======================================================
    if excel_file.exists():
        excel_file.unlink()
    wb.save(excel_file)
