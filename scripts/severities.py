from axonius_api_client import Connect
import os
from pathlib import Path
from dotenv import load_dotenv
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from collections import defaultdict
import datetime as dt
import warnings, axonius_api_client


def get_severities(severidad, central, f_central):
    current_date_and_time = str(dt.date.today())
    from table import mostrar_tabla
    from centrales  import centrales
    from critical import critical
    mostrar_tabla(centrales, current_date_and_time)
    warnings.filterwarnings(action="ignore", category=axonius_api_client.exceptions.ExtraAttributeWarning)
    warnings.filterwarnings("ignore")
    base_dir = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/'
    os.makedirs(base_dir, exist_ok=True)
    """if central == "IXTLA":
        done_path = os.path.join(base_dir, "done", f"{severidad.lower()}_{central}.done")
        os.makedirs(os.path.dirname(done_path), exist_ok=True)
        with open(done_path, "w") as f:
            f.write("done")
        mostrar_tabla(centrales, current_date_and_time)
        return"""

    dotenv_path = Path(__file__).parent / ".env"
    load_dotenv(dotenv_path=dotenv_path)

    connect_args = {
        "url": os.getenv("AXONIUS_URL"),
        "key": os.getenv("AXONIUS_KEY"),
        "secret": os.getenv("AXONIUS_SECRET"),
        "verify": False
    }

    client = Connect(**connect_args)
    apiobj = client.devices
    #query_name = f"Servers in {f_central} Vuln {severidad}"
    query_name = f"{severidad} VULNERABILITIES SERVERS IN {f_central}"
    try:
        devices = apiobj.get_by_saved_query(query_name)
    except Exception:
        print(f"Query NOT FOUND: {query_name}. LAST RELEASE OF CRITICAL...")
        critical(central=central, current_date_and_time=current_date_and_time, severidad=severidad)
        return  # detener ejecución

    mostrar_tabla(centrales, current_date_and_time)
    # Crear carpeta de salida
    
    base_dir = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/'
    os.makedirs(base_dir, exist_ok=True)


    if not devices or len(devices) == 0:
        print(f"Empty query: {query_name}.")
        done_path = os.path.join(base_dir, "done", f"{severidad.lower()}_{central}.done")
        os.makedirs(os.path.dirname(done_path), exist_ok=True)
        with open(done_path, "w") as f:
            f.write("done")
        return

    # Guardar JSON dentro de la carpeta destino
    json_path = os.path.join(base_dir, f"{severidad}_SEV_{central}.json")

    filtered_devices = []
    for device in devices:
        filtered_device = {
            "adapters": device.get("adapters", []),
            "hostname": device.get("specific_data.data.hostname_preferred"),
            "ips": device.get("specific_data.data.network_interfaces.ips_preferred", []),
            "macs": device.get("specific_data.data.network_interfaces.mac_preferred", []),
            "os": device.get("specific_data.data.os.type_distribution_preferred"),
            "software_cves": []
        }

        for cve in device.get("specific_data.data.software_cves", []):
            filtered_cve = {
                "cve_description": cve.get("cve_description"),
                "cve_id": cve.get("cve_id"),
                "cve_severity": cve.get("cve_severity"),
                "software_name": cve.get("software_name"),
                "software_version": cve.get("software_version"),
            }
            filtered_device["software_cves"].append(filtered_cve)

        filtered_devices.append(filtered_device)

    # Si no hay datos, no generar nada
    if not filtered_devices:
        return

    # Guardar JSON
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(filtered_devices, f, indent=4, ensure_ascii=False)

    # Cargar JSON
    with open(json_path, "r", encoding="utf-8") as f:
        devices = json.load(f)

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"{severidad}_SEV_{central}_{current_date_and_time}"
    MAX_EXCEL_ROWS = 1048576
    current_row = 1  # encabezado ya ocupa la fila 1
    excel_limit_reached = False

    headers = [
        "Adaptadores", "CVE", "Numero de Dispositivos", "Severidad", "Descripcion",
        "Hostname", "IPs", "MAC", "Tipo y distribución OS", "Cortex", "Virtual Patching"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    cve_to_hostnames = defaultdict(set)
    cve_info = {}
    for device in devices:
        hostname = device.get("hostname", "")
        for cve in device.get("software_cves", []):
            cve_id = cve.get("cve_id")
            if cve_id:
                cve_to_hostnames[cve_id].add(hostname)
                cve_info[cve_id] = {"description": cve.get("cve_description", "")}

    for device in devices:
        adapters = "\n".join(device.get("adapters", []))
        hostname = device.get("hostname", "")
        ips = "\n".join(device.get("ips", []))
        macs = "\n".join(device.get("macs", []))
        os_name = device.get("os", "")
        cortex = "SI" if "paloalto_xdr_adapter" in device.get("adapters", []) else "NO"
        vpatch = "SI" if "deep_security_adapter" in device.get("adapters", []) else "NO"

        for cve in device.get("software_cves", []):
            if current_row >= MAX_EXCEL_ROWS:
                excel_limit_reached = True
                break

            cve_id = cve.get("cve_id", "")
            num_devices = len(cve_to_hostnames.get(cve_id, []))
            ws.append([
                adapters, cve_id, num_devices, cve.get("cve_severity", ""), cve.get("cve_description", ""),
                hostname, ips, macs, os_name, cortex, vpatch
            ])
            current_row += 1

        if excel_limit_reached:
            break


    # Ajustes visuales
    align_wrap = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    for col in ["A", "E", "G", "H"]:
        for cell in ws[col]:
            if cell.row > 1:
                cell.alignment = align_wrap

    for col, width in {
        "A": 30, "B": 20, "C": 25, "D": 10, "E": 40, "F": 30,
        "G": 25, "H": 25, "I": 30, "J": 15, "K": 20
    }.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = ws.dimensions

    # Hoja resumen
    ws_resumen = wb.create_sheet("RESUMEN")
    ws_resumen.append(["CVE", severidad, "Descripcion"])
    for cve_id, info in cve_info.items():
        num_devices = len(cve_to_hostnames.get(cve_id, []))
        ws_resumen.append([cve_id, num_devices, info.get("description", "")])

    for cell in ws_resumen[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col, width in {"A": 30, "B": 20, "C": 40}.items():
        ws_resumen.column_dimensions[col].width = width
    # Habilitar wrap text en la columna de Descripción (C)
    for cell in ws_resumen["C"]:
        if cell.row > 1:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        

    excel_path = os.path.join(base_dir, f"{severidad}_SEV_{central}_{current_date_and_time}.xlsx")
    wb.save(excel_path)

    # Crear archivo .done
    done_path = os.path.join(base_dir, "done", f"{severidad.lower()}_{central}.done")
    os.makedirs(os.path.dirname(done_path), exist_ok=True)
    with open(done_path, "w") as f:
        f.write("done")


    # Eliminar JSON temporal
    if os.path.exists(json_path):
        os.remove(json_path)
    mostrar_tabla(centrales, current_date_and_time)


def main():
    get_severities(severidad="CRITICAL", central="POLANCO", f_central="POLANCO")

if __name__ == "__main__":
    main()