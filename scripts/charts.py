import os
import datetime as dt
from openpyxl import load_workbook
import csv
import xlwings as xw
from centrales import centrales
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter


def sobrescribir_registro_csv(ruta_csv, fila_nueva):
    fecha_hoy = fila_nueva[0]
    registros = []
    encabezado = None

    if os.path.exists(ruta_csv):
        with open(ruta_csv, mode="r", newline="") as file:
            reader = csv.reader(file)
            registros = list(reader)

        if registros:
            encabezado = registros[0]
            datos = registros[1:]
        else:
            datos = []
    else:
        datos = []

    datos_filtrados = [fila for fila in datos if fila and fila[0] != fecha_hoy]
    datos_filtrados.append(fila_nueva)

    if not encabezado:
        encabezado = ["Date"] + [f"Value{i+1}" for i in range(len(fila_nueva)-1)]

    with open(ruta_csv, mode="w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(encabezado)
        writer.writerows(datos_filtrados)

def data_central(central):
    ruta_base = "ARCHIVOS_REPORTES"
    current_date_and_time = str(dt.date.today())
    nombre_archivo = f"Reporte_Discovery_{central}_{current_date_and_time}.xlsx"
    ruta_completa = os.path.join(ruta_base, central, current_date_and_time, nombre_archivo)

    carpeta_script = os.path.dirname(os.path.realpath(__file__))
    ruta_graficas_dir = os.path.abspath(os.path.join(carpeta_script, "..", "src", "Graficas"))
    ruta_csv = os.path.join(ruta_graficas_dir, f"data_{central}.csv")

    if not os.path.exists(ruta_graficas_dir):
        os.makedirs(ruta_graficas_dir)
    wb = xw.Book(ruta_completa)
    wb.app.calculate()
    wb.save()
    wb.close()
    if os.path.exists(ruta_completa):
        try:
            wb_origen = load_workbook(ruta_completa, data_only=True)
            if "Resumen" in wb_origen.sheetnames:
                hoja = wb_origen["Resumen"]
                assets = hoja["E3"].value
                identified_dev = hoja["E4"].value
                servers_t = hoja["E5"].value
                servers_c = hoja["E6"].value
                servers_noc = str(int(hoja["E8"].value) + int(hoja["E7"].value))
                unidentified_serv = hoja["E14"].value
                unmanagged = hoja["E15"].value
                wb_origen.close()
            else:
                print("La hoja 'Resumen' no existe.")
                return None
        except Exception as e:
            print(f"Error al abrir el archivo origen: {e}")
            return None

        try:
            fecha_formateada = dt.date.today().strftime("%d/%m/%y")
            fila_nueva = [fecha_formateada, assets, identified_dev, servers_c, servers_noc, unidentified_serv, unmanagged,servers_t]
            sobrescribir_registro_csv(ruta_csv, fila_nueva)
        except Exception as e:
            print(f"Error al escribir en el CSV: {e}")
    else:
        print(f"El archivo origen no existe: {ruta_completa}")
        return False

def agregar_hojas_graficas(central):
    fecha_hoy = dt.date.today().strftime("%Y-%m-%d")
    ruta_reporte = os.path.join("ARCHIVOS_REPORTES", central, fecha_hoy, f"Reporte_Discovery_{central}_{fecha_hoy}.xlsx")

    carpeta_script = os.path.dirname(os.path.realpath(__file__))
    ruta_csv = os.path.abspath(os.path.join(carpeta_script, "..", "src", "Graficas", f"data_{central}.csv"))

    if not os.path.exists(ruta_reporte):
        print(f"No se encontró el archivo de reporte para {central}.")
        return

    try:
        wb = load_workbook(ruta_reporte)

        def crear_hoja_con_grafica(wb, nombre_hoja, ruta_csv, columnas_numericas):
            if nombre_hoja in wb.sheetnames:
                del wb[nombre_hoja]
            ws = wb.create_sheet(nombre_hoja)

            with open(ruta_csv, newline='') as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader, start=1):
                    for col_idx, valor in enumerate(row, start=1):
                        if row_idx > 1 and col_idx in columnas_numericas:
                            try:
                                valor = float(valor.replace(",", "")) if isinstance(valor, str) else float(valor)
                            except:
                                pass
                        ws.cell(row=row_idx, column=col_idx).value = valor

            chart = LineChart()
            chart.title = f"Resumen {central}"
            chart.style = 2
            chart.y_axis.title = "Cantidad"
            chart.x_axis.title = "Fecha"

            # Ampliar tamaño visual
            chart.width = 38
            chart.height = 20

            max_row = ws.max_row
            for col_idx in columnas_numericas:
                data = Reference(ws, min_col=col_idx, min_row=1, max_row=max_row)
                chart.add_data(data, titles_from_data=True)

            cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            chart.set_categories(cats)

            col_chart = get_column_letter(max(columnas_numericas) + 2)
            ws.add_chart(chart, f"{col_chart}2")

        crear_hoja_con_grafica(wb, f"Data_{central}", ruta_csv, [2,3,4,5,6,7,8])
        wb.save(ruta_reporte)
    except Exception as e:
        print(f"Error al crear hojas con gráficas: {e}")

if __name__ == '__main__':
    for central in centrales:
        try:
            data_central(central.nombre)
            agregar_hojas_graficas(central.nombre)
        except Exception as e:
            print(f"Error : {e}")
    #data_central("IXTLA")
    #agregar_hojas_graficas("IXTLA")