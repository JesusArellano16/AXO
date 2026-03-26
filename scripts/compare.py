import json
from pathlib import Path
from openpyxl import Workbook
import sys
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent

CATEGORY_MAP = {
    "ALL_GENERAL_NETWORK_DEVICES": "ALL_GENERAL_NETWORK_DEVICES.json",
    "ALL_GENERAL_UNIDENTIFIED_SERVERS": "ALL_GENERAL_UNIDENTIFIED_SERVERS.json",
    "GENERAL_ASSETS": "GENERAL_ASSETS.json",
    "GENERAL_PCs": "GENERAL_PCs.json",
    "GENERAL_SERVERS": "GENERAL_SERVERS.json",
    "GENERAL_VARIOUS_IDENTIFIED_DEVICES": "GENERAL_VARIOUS_IDENTIFIED_DEVICES.json"
}


def load_json(fecha, filename):
    path = BASE_DIR / "RECORD/GENERAL_JSON" / fecha / filename
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def get_region_labels(region):
    unique_labels_path = BASE_DIR / "scripts" / "UNIQUE_LABELS.json"
    with open(unique_labels_path, "r", encoding="utf-8") as f:
        unique_labels = json.load(f)["unique_labels"]

    region_number = int(region.replace("R", ""))
    return {
        label for label, r in unique_labels.items()
        if r == region_number
    }


def filter_assets(assets, central):
    if central == "GENERAL":
        return assets

    if central.startswith("R"):
        region_labels = get_region_labels(central)
        return [
            a for a in assets
            if any(l in region_labels for l in a.get("labels", []))
        ]

    return [
        a for a in assets
        if central in a.get("labels", [])
    ]


def normalize_asset(asset):
    return {
        "id": asset.get("internal_axon_id"),
        "hostname": asset.get("specific_data.data.hostname_preferred"),
        "ips": sorted(asset.get("specific_data.data.network_interfaces.ips_preferred", [])),
        "macs": sorted(asset.get("specific_data.data.network_interfaces.mac_preferred", [])),
        "adapters": sorted(list(set(asset.get("adapters", [])))),
        "os": asset.get("specific_data.data.os.type_distribution_preferred")
    }


def compare_assets(ref_assets, comp_assets):
    ref_by_id = {a["id"]: a for a in ref_assets if a["id"]}
    comp_by_id = {a["id"]: a for a in comp_assets if a["id"]}

    ref_by_host = {a["hostname"]: a for a in ref_assets if a["hostname"]}
    comp_by_host = {a["hostname"]: a for a in comp_assets if a["hostname"]}

    visited_comp_ids = set()
    results = []

    for rid, ref in ref_by_id.items():
        comp = comp_by_id.get(rid)

        if not comp and ref.get("hostname"):
            comp = comp_by_host.get(ref["hostname"])

        if comp:
            if comp.get("id"):
                visited_comp_ids.add(comp["id"])

            changes = []

            if ref.get("ips") != comp.get("ips"):
                changes.append("Modified IPs")

            if ref.get("hostname") != comp.get("hostname"):
                changes.append("Modified Hostname")

            if ref.get("adapters") != comp.get("adapters"):
                changes.append("Modified Adapters")

            if ref.get("os") != comp.get("os"):
                changes.append("Modified OS")

            if not changes:
                status = "UNCHANGED"
                comment = ""
            else:
                status = "MODIFIED"
                comment = ", ".join(changes)

        else:
            status = "REMOVED"
            comment = ""

        results.append({
            "id": ref.get("id"),
            "hostname": ref.get("hostname"),
            "ref": ref,
            "comp": comp,
            "status": status,
            "comment": comment
        })

    for cid, comp in comp_by_id.items():
        if cid in visited_comp_ids:
            continue

        ref = None
        if comp.get("hostname"):
            ref = ref_by_host.get(comp["hostname"])

        if ref:
            continue

        results.append({
            "id": comp.get("id"),
            "hostname": comp.get("hostname"),
            "ref": None,
            "comp": comp,
            "status": "NEW",
            "comment": ""
        })

    return results

def generate_excel(results, output_file, fecha1, fecha2, category, total_ref, total_comp):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Alignment, Font

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_view.zoomScale = 200

    ws_summary.column_dimensions["A"].width = 55

    bold = Font(bold=True)
    big_bold = Font(bold=True, size=14)
    red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws_summary["A1"] = "ASSET COMPARISON SUMMARY"
    ws_summary["A1"].font = big_bold

    ws_summary["A3"] = f"{fecha1} {category} (Reference)"
    ws_summary["B3"] = total_ref
    

    ws_summary["A4"] = f"{fecha2} {category} (Comparison)"
    ws_summary["B4"] = total_comp

    diff = total_comp - total_ref

    if diff > 0:
        ws_summary["D3"] = f"↑ {diff}"
        ws_summary["D3"].font = Font(bold=True, color="008000")  # verde
    elif diff < 0:
        ws_summary["D3"] = f"↓ {abs(diff)}"
        ws_summary["D3"].font = Font(bold=True, color="FF0000")  # rojo
    else:
        ws_summary["D3"] = "→ 0"
        ws_summary["D3"].font = Font(bold=True)
    new_count = sum(1 for r in results if r["status"] == "NEW")
    removed_count = sum(1 for r in results if r["status"] == "REMOVED")
    modified = [r for r in results if r["status"] == "MODIFIED"]

    ws_summary["A6"] = "New Assets"
    ws_summary["B6"] = new_count
    ws_summary["B6"].fill = green

    ws_summary["A7"] = "Removed Assets"
    ws_summary["B7"] = removed_count
    ws_summary["B7"].fill = red

    ws_summary["A8"] = "Modified Assets"
    ws_summary["B8"] = len(modified)
    ws_summary["B8"].fill = yellow

    mod_ips = sum(1 for r in modified if "IPs" in r.get("comment", ""))
    mod_host = sum(1 for r in modified if "Hostname" in r.get("comment", ""))
    mod_adapters = sum(1 for r in modified if "Adapters" in r.get("comment", ""))
    mod_os = sum(1 for r in modified if "OS" in r.get("comment", ""))

    ws_summary["A10"] = "Modified IPs"
    ws_summary["B10"] = mod_ips

    ws_summary["A11"] = "Modified Hostname"
    ws_summary["B11"] = mod_host

    ws_summary["A12"] = "Modified Adapters"
    ws_summary["B12"] = mod_adapters

    ws_summary["A13"] = "Modified OS Type/Distribution"
    ws_summary["B13"] = mod_os

    ws_summary["A15"] = "VALIDATE INFORMATION"
    ws_summary["A15"].font = Font(bold=True, color="FF0000", size=14)

    ws = wb.create_sheet("Compare")

    headers = [
        "ID",
        "Comment",
        "Hostname Before",
        "Hostname After",
        "Status",
        "IPs Before",
        "IPs After",
        "Adapters Before",
        "Adapters After",
        "OS Before",
        "OS After"
    ]

    ws.append(headers)

    ws.auto_filter.ref = f"A1:K1"

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 25
    ws.column_dimensions["K"].width = 25

    wrap = Alignment(wrap_text=True)

    for r in results:
        ref = r["ref"] or {}
        comp = r["comp"] or {}

        ips_before = "\n".join(ref.get("ips", []))
        ips_after = "\n".join(comp.get("ips", []))

        adapters_before = "\n".join(ref.get("adapters", []))
        adapters_after = "\n".join(comp.get("adapters", []))

        row = [
            r["id"],
            r.get("comment", ""),
            ref.get("hostname"),
            comp.get("hostname"),
            r["status"],
            ips_before,
            ips_after,
            adapters_before,
            adapters_after,
            ref.get("os"),
            comp.get("os")
        ]

        ws.append(row)
        excel_row = ws.max_row

        for col in [6, 7, 8, 9]:
            ws.cell(row=excel_row, column=col).alignment = wrap

        if r["status"] == "NEW":
            fill = green
        elif r["status"] == "REMOVED":
            fill = red
        elif r["status"] == "MODIFIED":
            fill = yellow
        else:
            fill = None

        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=col).fill = fill

    wb.save(output_file)


def main():

    central, category, fecha1, fecha2 = sys.argv[1:]

    # 🔁 Normalización de central
    if central == "IXTLA":
        central = "IXTLAHUACA"

    filename = CATEGORY_MAP.get(category)

    ref_raw = load_json(fecha1, filename)
    comp_raw = load_json(fecha2, filename)

    ref_filtered = filter_assets(ref_raw, central)
    comp_filtered = filter_assets(comp_raw, central)

    ref_norm = [normalize_asset(a) for a in ref_filtered]
    comp_norm = [normalize_asset(a) for a in comp_filtered]

    results = compare_assets(ref_norm, comp_norm)
    ref_total = len(ref_norm)
    comp_total = len(comp_norm)

    output_dir = BASE_DIR / "Comparision_tool" / central / category
    output_dir.mkdir(parents=True, exist_ok=True)

    file_name = f"compare_{central}_{fecha1}_vs_{fecha2}.xlsx"
    output_file = output_dir / file_name

    if output_file.exists():
        output_file.unlink()

    generate_excel(
        results,
        output_file,
        fecha1,
        fecha2,
        category,
        ref_total,
        comp_total
    )

    print(f"Reporte generado: {output_file}")


if __name__ == "__main__":
    main()