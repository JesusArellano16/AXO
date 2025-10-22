import pandas as pd
import shutil
import csv
import os
import openpyxl
from openpyxl.styles import Alignment

def verificar_archivo(vuln, central):
    """
    Verifica si el archivo de vulnerabilidades existe en la carpeta de la central.
    """
    ruta_base = "AXONIUS_FILES"
    ruta_completa = os.path.join(ruta_base, central)
    
    if not os.path.isdir(ruta_completa):
        return False  # La carpeta no existe
    return f"{vuln}_{central}.csv" in os.listdir(ruta_completa)

def critical(central, current_date_and_time, severidad):
    #print(f'🚀 Iniciando proceso para {severidad} en {central} 0/9')
    from table import mostrar_tabla
    from centrales  import centrales
    mostrar_tabla(centrales, current_date_and_time)
    if not verificar_archivo(vuln=severidad.lower(), central=central):
        vuln=severidad.lower()
        #print(f'❌ {central} no tiene vulnerabilidades {severidad}')
        done_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/done/{vuln}_{central}.done'
        with open(done_path, 'w') as f:
            f.write("done")
        mostrar_tabla(centrales, current_date_and_time)
        return

    # Copiar archivo CSV a la carpeta de reportes
    #print(f'📂 Copiando archivo {severidad}.csv 1/9 en {central}')
    src_path = f'./AXONIUS_FILES/{central}/{severidad}_{central}.csv'
    dest_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/{severidad}.csv'
    shutil.copy(src_path, dest_path)
    
    # Leer el archivo CSV
    #print(f'📖 Leyendo archivo CSV... 2/9 en {central}')
    with open(src_path, encoding="utf-8") as file:
        csv_reader = csv.reader(file, delimiter=',')
        vulnerabilities = []
        devices = []
        for row in csv_reader:
            if row[0] == "Device":
                devices.append(row)
            elif row[0] == "Vulnerability":
                vulnerabilities.append(row)
    
    #print(f'✂️ Procesando datos de vulnerabilidades y dispositivos... 3/9 en {central}')
    for col in vulnerabilities:
        del col[5:]
        del col[0]
    for col in devices:
        del col[:5]
    
    # Crear archivo Excel
    #print(f'📊 Creando archivo Excel... 4/9 en {central}')
    
    namew = f'{severidad.upper()}_SEV_{central}_{current_date_and_time}'
    name = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/{namew}.xlsx'
    
    try:
        read_file_product = pd.read_csv(dest_path, on_bad_lines="skip")
    except pd.errors.ParserError as e:
        print(f"❌ Error leyendo {dest_path}: {e}")
        return
    read_file_product.to_excel(name, index=None, header=True)
    os.remove(dest_path)
    
    wb = openpyxl.load_workbook(name)
    ws = wb['Sheet1']
    ws.title = namew
    
    #print(f'📑 Creando hoja CVE... 5/9 en {central}')
    wb.create_sheet('CVE')
    ws = wb['CVE']
    ws.append(["Adaptadores", "CVE", "Device Count", "Severity", "Description", "Adaptadores"])
    for vul in vulnerabilities:
        vul.insert(3, severidad.upper())
        vul.append(vul[0])
        ws.append(vul)
    columnas_objetivo = [1, 5, 6]
    columnas_ancho = {
            "A": 30,
            "B": 20,
            "C": 15,
            "D": 10,
            "E": 40,
            "F": 30
        }
    fil = "A1:F1"
    for col in columnas_objetivo:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)  # Activar "Wrap Text"

    for col, width in columnas_ancho.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = fil
    #print(f'🖥️ Creando hoja de dispositivos... 6/9 en {central}')
    wb.remove(wb[namew])
    wb.create_sheet(namew)
    ws = wb[namew]
    ws.append(["Adaptadores", "CVE", "Numero de Dispositivos", "Severidad", "Descripcion", "Hostname", "IPs", "MAC", "Tipo y distribución OS", "Cortex", "Virtual Patching"])
    
    for dev in devices:
        dev.insert(0, dev[5])
        dev.pop()
        dev.insert(0, dev[5])
        dev.pop()
        for vul in vulnerabilities:
            if dev[1] == vul[1]:
                dev.insert(2, vul[2])
                dev.insert(3, vul[3])
                dev.insert(4, vul[4])
        dev.append("SI" if 'paloalto' in dev[0] else "NO")
        dev.append("SI" if 'deep_security_adapter' in dev[0] else "NO")
        ws.append(dev)

    columnas_objetivo = [1, 5, 7,8]
    columnas_ancho = {
            "A": 30,
            "B": 20,
            "C": 15,
            "D": 10,
            "E": 40,
            "F": 50,
            "G": 20,
            "H": 20,
            "I": 25,
            "J": 8,
            "K": 15
        }
    fil = "A1:K1"
    for col in columnas_objetivo:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)  # Activar "Wrap Text"

    for col, width in columnas_ancho.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = fil
    
    wb.save(name)
    
    #print(f'📄 Generando resumen en CSV... 7/9 en {central}')
    csv_file_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/example.csv'
    with open(csv_file_path, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Adaptadores", "CVE", "Numero de Dispositivos", "Severidad", "Descripcion", "Hostname", "IPs", "MAC", "Tipo y distribución OS", "Cortex","Virtual Patching"])
        writer.writerows(devices)
    
    try:
        df_devides = pd.read_csv(csv_file_path, encoding='utf-8', delimiter=',', on_bad_lines="skip")
    except pd.errors.ParserError as e:
        print(f"❌ Error leyendo {csv_file_path}: {e}")
        return
    de = df_devides.pivot_table(index="CVE", columns="Severidad", values="Numero de Dispositivos")
    
    with pd.ExcelWriter(name, engine="openpyxl", mode='a') as writer:
        de.to_excel(writer, sheet_name="RESUMEN")
    
    try:
        os.remove(csv_file_path)
    except:
        pass
    
    #print(f'✍️ Aplicando formato en el Excel... 8/9 en {central}')
    wb = openpyxl.load_workbook(name)

    ws = wb['RESUMEN']
    columnas_ancho = {
            "A": 30,
            "B": 20
        }
    fil = "A1:B1"

    for col, width in columnas_ancho.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = fil
        
    wb.save(name)
    vuln=severidad.lower()
        #print(f'❌ {central} no tiene vulnerabilidades {severidad}')
    done_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/done/{vuln}_{central}.done'
    with open(done_path, 'w') as f:
        f.write("done")
    mostrar_tabla(centrales, current_date_and_time)
    return
    #print(f'✅ Proceso finalizado: {name} creado exitosamente 9/9')
