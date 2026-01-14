from axonius_api_client import Connect
import os
from pathlib import Path
from dotenv import load_dotenv
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import datetime as dt
import warnings, axonius_api_client


# ======================================================
# UTILIDAD JSON OFFLINE
# ======================================================
def load_eol_from_json(central):
    base_dir = Path(__file__).parent.parent / "AXONIUS_FILES" / "GENERAL_JSON"
    json_path = base_dir / "EoL_GENERAL_SERVERS.json"

    if not json_path.exists():
        raise FileNotFoundError(json_path)

    with open(json_path, "r", encoding="utf-8") as f:
        devices = json.load(f)

    # Filtrar por label
    return [d for d in devices if central in d.get("labels", [])]


# ======================================================
# FUNCIÓN PRINCIPAL
# ======================================================
def export_eol(central, f_central):
    current_date_and_time = str(dt.date.today())

    from table import mostrar_tabla
    from centrales import centrales
    from eol import Eol

    mostrar_tabla(centrales, current_date_and_time)

    warnings.filterwarnings("ignore", category=axonius_api_client.exceptions.ExtraAttributeWarning)
    warnings.filterwarnings("ignore")

    # --------------------------------------------------
    # PATHS
    # --------------------------------------------------
    base_dir = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/'
    os.makedirs(f'{base_dir}/done', exist_ok=True)

    # --------------------------------------------------
    # DATA SOURCE LOGIC
    # --------------------------------------------------
    if central == "IXTLA":
        dotenv_path = Path(__file__).parent / ".env"
        load_dotenv(dotenv_path=dotenv_path)

        connect_args = {
            "url": os.getenv("AXONIUS_URL"),
            "key": os.getenv("AXONIUS_KEY"),
            "secret": os.getenv("AXONIUS_SECRET"),
            "verify": False
        }

        client = Connect(**connect_args)
        apiobj = client.devices

        query_name = f"DEV EoL SERVERS IN {f_central}"

        try:
            devices = apiobj.get_by_saved_query(
                query_name,
                fields=[
                    "labels",
                    "adapters",
                    "specific_data.data.hostname_preferred",
                    "specific_data.data.installed_software.preferred_name",
                    "specific_data.data.installed_software.version",
                    "specific_data.data.installed_software.end_of_life",
                    "specific_data.data.installed_software.end_of_support_date",
                    "specific_data.data.network_interfaces.ips_preferred",
                    "specific_data.data.network_interfaces.mac_preferred",
                    "specific_data.data.os.type_distribution_preferred",
                ],
            )
        except Exception:
            print(f"Query NOT FOUND: {query_name}. LAST RELEASE OF EOL...")
            Eol(central=central, current_date_and_time=current_date_and_time)
            return
    else:
        devices = load_eol_from_json(central)

    mostrar_tabla(centrales, current_date_and_time)

    if not devices:
        done_path = os.path.join(base_dir, "done", f"eol_{central.lower()}.done")
        with open(done_path, "w") as f:
            f.write("done")
        return

    # --------------------------------------------------
    # EXCEL
    # --------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = f"EOL_{central}_{current_date_and_time}"

    headers = [
        "Adaptadores",
        "Preferred Host Name",
        "Installed Software",
        "Software Version",
        "End of Life",
        "End Of Support",
        "IPs",
        "MAC",
        "Tipo y distribución OS",
        "Cortex",
        "Virtual Patching",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --------------------------------------------------
    # DATA
    # --------------------------------------------------
    for device in devices:
        adapters = "\n".join(device.get("adapters", []))
        hostname = device.get("specific_data.data.hostname_preferred", "")
        ips = "\n".join(device.get("specific_data.data.network_interfaces.ips_preferred", []))
        macs = "\n".join(device.get("specific_data.data.network_interfaces.mac_preferred", []))
        os_name = device.get("specific_data.data.os.type_distribution_preferred", "")
        cortex = "SI" if "paloalto_xdr_adapter" in device.get("adapters", []) else "NO"
        vpatch = "SI" if "deep_security_adapter" in device.get("adapters", []) else "NO"

        softwares = device.get("specific_data.data.installed_software", [])

        if not softwares:
            ws.append([adapters, hostname, "", "", "", "", ips, macs, os_name, cortex, vpatch])
        else:
            for sw in softwares:
                ws.append([
                    adapters,
                    hostname,
                    sw.get("preferred_name", ""),
                    sw.get("version", ""),
                    sw.get("end_of_life", ""),
                    sw.get("end_of_support_date", ""),
                    ips,
                    macs,
                    os_name,
                    cortex,
                    vpatch
                ])

    # --------------------------------------------------
    # FORMATO (MISMO QUE LOS OTROS REPORTES)
    # --------------------------------------------------
    wrap_columns = ["A", "C", "G", "H"]

    for col in wrap_columns:
        for cell in ws[col]:
            if cell.row > 1:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="bottom",
                    wrap_text=True
                )

    for col, width in {
        "A": 30, "B": 30, "C": 35, "D": 20, "E": 20, "F": 20,
        "G": 25, "H": 25, "I": 30, "J": 15, "K": 20
    }.items():
        ws.column_dimensions[col].width = width

    ws.auto_filter.ref = ws.dimensions

    # --------------------------------------------------
    # SAVE
    # --------------------------------------------------
    excel_path = os.path.join(base_dir, f"EOL_{central}_{current_date_and_time}.xlsx")
    wb.save(excel_path)

    done_path = os.path.join(base_dir, "done", f"eol_{central.lower()}.done")
    with open(done_path, "w") as f:
        f.write("done")

    mostrar_tabla(centrales, current_date_and_time)


# ======================================================
# MAIN (PRUEBA)
# ======================================================
def main():
    export_eol(central="CARSO", f_central="CARSO")


if __name__ == "__main__":
    main()
