import sys
import pandas as pd
import shutil
import csv
import os
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

def verificar_archivo(vuln, central):
    ruta_base = "AXONIUS_FILES"
    ruta_completa = os.path.join(ruta_base, central)
    if not os.path.isdir(ruta_completa):
        return False
    return f"{vuln}_{central}.csv" in os.listdir(ruta_completa)

def critical(central, current_date_and_time, severidad):

    try:
        vuln = severidad.lower()
        if not verificar_archivo(vuln=vuln, central=central):
            done_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/done/{vuln}_{central}.done'
            os.makedirs(os.path.dirname(done_path), exist_ok=True)
            with open(done_path, 'w') as f:
                f.write("done")
            return

        src_path = f'./AXONIUS_FILES/{central}/{severidad}_{central}.csv'
        dest_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/{severidad}.csv'
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        shutil.copy(src_path, dest_path)

        vulnerabilities, devices = [], []
        with open(src_path, encoding="utf-8") as file:
            csv_reader = csv.reader(file, delimiter=',')
            for i, row in enumerate(csv_reader, start=1):
                try:
                    if not row:
                        continue
                    if row[0] == "Device":
                        devices.append(row)
                    elif row[0] == "Vulnerability":
                        vulnerabilities.append(row)
                except Exception as e:
                    print(f"Error en fila {i}: {row}")
                    print(f"   -> {type(e).__name__}: {e}")


        try:
            for col in vulnerabilities:
                del col[5:]
                del col[0]
            for col in devices:
                del col[:5]
        except Exception as e:
            print(f"Error procesando columnas: {type(e).__name__}: {e}")
            raise

        namew = f'{severidad.upper()}_SEV_{central}_{current_date_and_time}'
        name = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/{namew}.xlsx'
        try:
            read_file_product = pd.read_csv(dest_path, on_bad_lines="skip")
            read_file_product.to_excel(name, index=None, header=True)
            os.remove(dest_path)
        except Exception as e:
            raise

        wb = openpyxl.load_workbook(name)
        ws = wb.active
        ws.title = namew
        wb.create_sheet('CVE')
        ws = wb['CVE']
        ws.append(["Adaptadores", "CVE", "Device Count", "Severity", "Description", "Adaptadores"])
        for vul in vulnerabilities:
            vul.insert(3, severidad.upper())
            vul.append(vul[0])
            ws.append(vul)

        for col, width in {"A": 30, "B": 20, "C": 15, "D": 10, "E": 40, "F": 30}.items():
            ws.column_dimensions[col].width = width
        ws.auto_filter.ref = "A1:F1"

        if namew in wb.sheetnames:
            wb.remove(wb[namew])
        wb.create_sheet(namew)
        ws = wb[namew]
        ws.append(["Adaptadores", "CVE", "Numero de Dispositivos", "Severidad", "Descripcion",
                   "Hostname", "IPs", "MAC", "Tipo y distribución OS", "Cortex", "Virtual Patching"])
        valid_devices = devices.copy()
        filtered_devices = []
        for i, dev in enumerate(valid_devices, start=1):
            if len(dev) == 0 or dev[-1].strip() == "":
                continue
            try:
                dev.insert(0, dev[5])
                dev.pop()
                dev.insert(0, dev[5])
                dev.pop()
                for vul in vulnerabilities:
                    if dev[1] == vul[1]:
                        dev.insert(2, vul[2])
                        dev.insert(3, vul[3])
                        dev.insert(4, vul[4])
                dev.append("SI" if 'paloalto' in dev[0] else "NO")
                dev.append("SI" if 'deep_security_adapter' in dev[0] else "NO")
                ws.append(dev)
                filtered_devices.append(dev)
            except Exception as e:
                print(f"Error procesando dispositivo #{i}: {dev}")
                print(f"   -> {type(e).__name__}: {e}")
        devices = filtered_devices

        for col, width in {"A": 30, "B": 20, "C": 15, "D": 10, "E": 40, "F": 50, "G": 20,
                           "H": 20, "I": 25, "J": 8, "K": 15}.items():
            ws.column_dimensions[col].width = width
        ws.auto_filter.ref = "A1:K1"

        wb.save(name)

        csv_file_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/example.csv'
        with open(csv_file_path, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["Adaptadores", "CVE", "Numero de Dispositivos", "Severidad",
                             "Descripcion", "Hostname", "IPs", "MAC", "Tipo y distribución OS",
                             "Cortex", "Virtual Patching"])
            writer.writerows(devices)

        df_devices = pd.read_csv(csv_file_path, encoding='utf-8', delimiter=',', on_bad_lines="skip")
        de = df_devices.pivot_table(index="CVE", columns="Severidad", values="Numero de Dispositivos")

        with pd.ExcelWriter(name, engine="openpyxl", mode='a') as writer:
            de.to_excel(writer, sheet_name="RESUMEN")

        if os.path.exists(csv_file_path):
            os.remove(csv_file_path)

        wb = openpyxl.load_workbook(name)
        ws = wb['RESUMEN']
        for col, width in {"A": 30, "B": 20}.items():
            ws.column_dimensions[col].width = width
        ws.auto_filter.ref = "A1:B1"
        wb.save(name)

        done_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/done/{vuln}_{central}.done'
        os.makedirs(os.path.dirname(done_path), exist_ok=True)
        with open(done_path, 'w') as f:
            f.write("done")


    except Exception as e:
        print(f"\nError general: {type(e).__name__}: {e}\n")

def main():
    if len(sys.argv) < 4:
        sys.exit(1)

    central = sys.argv[1]
    current_date_and_time = sys.argv[2]
    severidad = sys.argv[3]
    critical(central, current_date_and_time, severidad)

if __name__ == "__main__":
    main()
