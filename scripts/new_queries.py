import axonius_api_client as axonapi  # Importar la librería de Axonius API Client
import json  # Importar la librería JSON para manejar datos en formato JSON
import os
import urllib3
import warnings  # Importar warnings para manejar advertencias
from dotenv import load_dotenv
from pathlib import Path
from openpyxl import load_workbook
from openpyxl import styles
import datetime as dt
from classifications import subClassification
from openpyxl.styles import Alignment
from copy import copy

dotenv_path = Path(__file__).parent / ".env"
load_dotenv(dotenv_path=dotenv_path)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.simplefilter("ignore",category=Warning)  # Ignorar advertencias para evitar mensajes innecesarios en la consola


connect_args = {
    "url": os.getenv("AXONIUS_URL"),  # IP de la instancia de Axonius
    "key": os.getenv("AXONIUS_KEY"),  # API Key proporcionada
    "secret": os.getenv("AXONIUS_SECRET"),  # API Secret proporcionado
    "verify": False  # Desactivar la verificación SSL (No recomendado en producción)
}
columns = [
    "adapters",
    "specific_data.data.network_interfaces.ips_preferred",
    "specific_data.data.network_interfaces.mac_preferred",
    "specific_data.data.network_interfaces.manufacturer",
    "Clasificacion"
]
#central = "IXTLA"

def copy_style_from_row(ws, source_row, start_row, end_row):
    for row in range(start_row, end_row + 1):
        for col in range(1, ws.max_column + 1):
            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=row, column=col)

            # Copiar estilos individuales
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)


def formater(ws):
    # Definir las columnas objetivo (A, B y C corresponden a índices 1, 2 y 3 en openpyxl)
    columnas_objetivo = [1, 2, 3]
    columnas_ancho = {
        "A": 25,
        "B": 20,
        "C": 25,
        "D": 50,
        "E": 70
    }
    fil = "A2:E2"


    # Aplicar ajuste de texto en las columnas A, C y D (desde la fila 2 en adelante)
    for col in columnas_objetivo:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)  # Activar "Wrap Text"

    for col, width in columnas_ancho.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = fil

def new_queries(central):
    if central == "IXTLA": central2 = "IXTLAHUACA"
    if central == "CARSO": central2 = central

    client = axonapi.Connect(**connect_args)
    saved_query_name = {
        f"ALL UNIDENTIFIED SERVERS {central2}":"Inventario No Identificados",
        f"VARIOUS IDENTIFIED DEVICES {central2}":"Inventario Identificados"
        }
    # Seleccionar el objeto API para dispositivos
    apiobj = client.devices

    for query_name, sheet_name in saved_query_name.items():
        # Obtener los dispositivos de la consulta guardada
        devices = apiobj.get_by_saved_query(query_name)
        for device in devices:  
            device.pop("adapter_list_length",None)
            device.pop("internal_axon_id",None)
            device.pop("labels",None)
            device.pop("specific_data.connection_label",None)
            # Normalizar manufacturer
            manufacturer = device.get("specific_data.data.network_interfaces.manufacturer", "")
            if isinstance(manufacturer, list) and manufacturer:
                device["specific_data.data.network_interfaces.manufacturer"] = manufacturer[0]
            #value_a = device.get("specific_data.data.network_interfaces.manufacturer")
            value_a = device.get("specific_data.data.network_interfaces.manufacturer")
            matching_key = next((k for k, v_list in subClassification.items() if value_a in v_list), "Desconocido")
            device["Clasificacion"] = matching_key
        with open(f"devices{central}.json", "w", encoding="utf-8") as f:
            json.dump(devices, f, indent=4, ensure_ascii=False)
        headers = [
            "Adaptadores",
            "IPs",
            "MACs",
            "Manufacturer",
            "Clasificacion"
        ]
        current_date_and_time = str(dt.date.today())
        path = r'./ARCHIVOS_REPORTES/' + central + r'/' + current_date_and_time + r'/' + f'REPORTE_DISCOVERY_{central}_{current_date_and_time}.xlsx'
        wb = load_workbook(path)
        ws = wb[sheet_name]


        # Limpiar el contenido anterior (desde la fila 2 en adelante)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.value = None

        fil = "A2:F2"
        ws.auto_filter.ref = fil

        # Escribir encabezados si es necesario
        ws['A2'].value = headers[0]
        ws['B2'].value = headers[1]
        ws['C2'].value = headers[2]
        ws['D2'].value = headers[3]
        ws['E2'].value = headers[4]

        row_num = 3
        # Escribir filas
        print(f"Escribiendo {len(devices)} dispositivos en la hoja '{sheet_name}'...")
        for device in devices:
            #row_num = ws.max_row + 1
            for col_num, key in enumerate(columns, 1):
                value = device.get(key, "")
                # Si el valor es una lista, convertir a string para que no dé error en Excel
                if isinstance(value, list):
                    value = "\n".join(str(v) for v in value)
                cell = ws.cell(row=row_num, column=col_num, value=value)
            row_num += 1
           # Guardar cambios
        copy_style_from_row(ws, source_row=3, start_row=2, end_row=row_num - 1)
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f"{col}2"].font = styles.Font(bold=True)
        formater(ws=ws)
        wb.save(path)
        wb.close()

