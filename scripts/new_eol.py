from axonius_api_client import Connect
import os
from pathlib import Path
from dotenv import load_dotenv
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import datetime as dt
import warnings, axonius_api_client

def export_eol(central, f_central):
    current_date_and_time = str(dt.date.today())
    from table import mostrar_tabla
    from centrales  import centrales
    from eol import Eol
    mostrar_tabla(centrales, current_date_and_time)
    warnings.filterwarnings(action="ignore", category=axonius_api_client.exceptions.ExtraAttributeWarning)
    warnings.filterwarnings("ignore")

    # Cargar credenciales
    dotenv_path = Path(__file__).parent / ".env"
    load_dotenv(dotenv_path=dotenv_path)

    connect_args = {
        "url": os.getenv("AXONIUS_URL"),
        "key": os.getenv("AXONIUS_KEY"),
        "secret": os.getenv("AXONIUS_SECRET"),
        "verify": False
    }

    client = Connect(**connect_args)
    apiobj = client.devices

    # 🔍 Query exacta en Axonius
    #query_name = f"AX - Servers in {f_central} with EOL Software"
    query_name = f"DEV EoL SERVERS IN {f_central}"
    try:
        devices = apiobj.get_by_saved_query(query_name)
    except Exception:
        print(f"Query NOT FOUND: {query_name}. LAST RELEASE OF EOL...")
        Eol(central=central, current_date_and_time=current_date_and_time)
        return  # detener ejecución

    # ✅ Incluir los campos necesarios para refine data
    devices = apiobj.get_by_saved_query(
        query_name,
        fields=[
            "adapters",
            "specific_data.data.hostname_preferred",
            "specific_data.data.installed_software.preferred_name",
            "specific_data.data.installed_software.version",
            "specific_data.data.installed_software.end_of_life",
            "specific_data.data.installed_software.end_of_support_date",
            "specific_data.data.network_interfaces.ips_preferred",
            "specific_data.data.network_interfaces.mac_preferred",
            "specific_data.data.os.type_distribution_preferred",
        ],
    )

    # 📁 Crear carpeta destino
    base_dir = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/'
    os.makedirs(base_dir, exist_ok=True)
    mostrar_tabla(centrales, current_date_and_time)

    if not devices or len(devices) == 0:
        print(f"Empty query: {query_name}.")
        done_path = os.path.join(base_dir, "done", f"eol_{central}.done")
        os.makedirs(os.path.dirname(done_path), exist_ok=True)
        with open(done_path, "w") as f:
            f.write("done")
        mostrar_tabla(centrales, current_date_and_time)
        return


    # 📄 Guardar JSON
    json_path = os.path.join(base_dir, f"EOL_{central}_{current_date_and_time}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(devices, f, indent=4, ensure_ascii=False)

    # 📄 Cargar JSON
    with open(json_path, "r", encoding="utf-8") as f:
        devices = json.load(f)

    # 📘 Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"EOL_{central}_{current_date_and_time}"

    headers = [
        "Adaptadores",
        "Preferred Host Name",
        "Installed Software",
        "Software Version",
        "End of Life",
        "End Of Support",
        "IPs",
        "MAC",
        "Tipo y distribución OS",
        "Cortex",
        "Virtual Patching",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 🧩 Procesar dispositivos
    for device in devices:
        adapters = "\n".join(device.get("adapters", []))
        hostname = device.get("specific_data.data.hostname_preferred", "")
        ips = "\n".join(device.get("specific_data.data.network_interfaces.ips_preferred", []))
        macs = "\n".join(device.get("specific_data.data.network_interfaces.mac_preferred", []))
        os_name = device.get("specific_data.data.os.type_distribution_preferred", "")
        cortex = "SI" if "paloalto_xdr_adapter" in device.get("adapters", []) else "NO"
        vpatch = "SI" if "deep_security_adapter" in device.get("adapters", []) else "NO"

        installed_softwares = device.get("specific_data.data.installed_software", [])
        if not installed_softwares:
            ws.append([adapters, hostname, "", "", "", "", ips, macs, os_name, cortex, vpatch])
        else:
            for sw in installed_softwares:
                sw_name = sw.get("preferred_name", "")
                sw_version = sw.get("version", "")
                sw_eol = sw.get("end_of_life", "")
                sw_eos = sw.get("end_of_support_date", "")
                ws.append([adapters, hostname, sw_name, sw_version, sw_eol, sw_eos, ips, macs, os_name, cortex, vpatch])

    # 🎨 Ajustes visuales
    align_wrap = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    for col in ["A", "C", "G", "H"]:
        for cell in ws[col]:
            if cell.row > 1:
                cell.alignment = align_wrap

    for col, width in {
        "A": 30, "B": 30, "C": 35, "D": 20, "E": 20, "F": 20,
        "G": 25, "H": 25, "I": 30, "J": 15, "K": 20
    }.items():
        ws.column_dimensions[col].width = width

    ws.auto_filter.ref = ws.dimensions

    # 💾 Guardar Excel
    excel_path = os.path.join(base_dir, f"EOL_{central}_{current_date_and_time}.xlsx")
    wb.save(excel_path)

    # ✅ Archivo .done
    done_path = os.path.join(base_dir, "done", f"eol_{central.lower()}.done")
    os.makedirs(os.path.dirname(done_path), exist_ok=True)
    with open(done_path, "w") as f:
        f.write("done")

    # 🧹 Eliminar JSON temporal
    if os.path.exists(json_path):
        os.remove(json_path)
    mostrar_tabla(centrales, current_date_and_time)


def main():
    export_eol(central="POLANCO",f_central="POLANCO")

if __name__ == "__main__":
    main()
