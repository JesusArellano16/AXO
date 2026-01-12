import axonius_api_client as axonapi
import json
import os
import urllib3
import warnings
from dotenv import load_dotenv
from pathlib import Path
from openpyxl import load_workbook, styles
import datetime as dt
from classifications import subClassification
from openpyxl.styles import Alignment
from copy import copy

# ======================================================
# ENV & WARNINGS
# ======================================================
dotenv_path = Path(__file__).parent / ".env"
load_dotenv(dotenv_path=dotenv_path)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.simplefilter("ignore", category=Warning)

connect_args = {
    "url": os.getenv("AXONIUS_URL"),
    "key": os.getenv("AXONIUS_KEY"),
    "secret": os.getenv("AXONIUS_SECRET"),
    "verify": False
}

columns = [
    "adapters",
    "specific_data.data.network_interfaces.ips_preferred",
    "specific_data.data.network_interfaces.mac_preferred",
    "specific_data.data.network_interfaces.manufacturer",
    "Clasificacion"
]

# ======================================================
# JSON OFFLINE UTILS
# ======================================================
def load_from_json(json_name, central):
    base_dir = Path(__file__).parent.parent / "AXONIUS_FILES" / "GENERAL_JSON"
    json_path = base_dir / json_name

    if not json_path.exists():
        raise FileNotFoundError(json_path)

    with open(json_path, "r", encoding="utf-8") as f:
        devices = json.load(f)

    return [d for d in devices if central in d.get("labels", [])]


# ======================================================
# EXCEL UTILS
# ======================================================
def copy_style_from_row(ws, source_row, start_row, end_row):
    for row in range(start_row, end_row + 1):
        for col in range(1, ws.max_column + 1):
            src = ws.cell(row=source_row, column=col)
            tgt = ws.cell(row=row, column=col)
            tgt.font = copy(src.font)
            tgt.border = copy(src.border)
            tgt.fill = copy(src.fill)
            tgt.number_format = copy(src.number_format)
            tgt.protection = copy(src.protection)
            tgt.alignment = copy(src.alignment)


def formater(ws):
    columnas_objetivo = [1, 2, 3]
    columnas_ancho = {
        "A": 25,
        "B": 20,
        "C": 25,
        "D": 50,
        "E": 70
    }

    for col in columnas_objetivo:
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    horizontal="left",
                    vertical="bottom"
                )

    for col, width in columnas_ancho.items():
        ws.column_dimensions[col].width = width

    ws.auto_filter.ref = "A4:E4"


# ======================================================
# MAIN LOGIC
# ======================================================
def new_queries(central, full_name):
    central2 = full_name
    current_date_and_time = str(dt.date.today())

    # Queries equivalentes
    saved_query_name = {
        "ALL_GENERAL_UNIDENTIFIED_SERVERS.json": "Inventario No Identificados",
        "GENERAL_VARIOUS_IDENTIFIED_DEVICES.json": "Inventario Identificados"
    }

    # API solo para IXTLA
    client = None
    apiobj = None
    if central == "IXTLA":
        client = axonapi.Connect(**connect_args)
        apiobj = client.devices

    for json_or_query, sheet_name in saved_query_name.items():

        # --------------------------------------------------
        # DATA SOURCE
        # --------------------------------------------------
        if central == "IXTLA":
            query_name = (
                f"ALL UNIDENTIFIED SERVERS {central2}"
                if "UNIDENTIFIED" in json_or_query
                else f"VARIOUS IDENTIFIED DEVICES {central2}"
            )
            devices = apiobj.get_by_saved_query(query_name)
        else:
            devices = load_from_json(json_or_query, central)

        # --------------------------------------------------
        # CLEAN + CLASSIFICATION
        # --------------------------------------------------
        for device in devices:
            device.pop("adapter_list_length", None)
            device.pop("internal_axon_id", None)
            device.pop("labels", None)
            device.pop("specific_data.connection_label", None)

            manufacturer = device.get("specific_data.data.network_interfaces.manufacturer", "")
            if isinstance(manufacturer, list) and manufacturer:
                manufacturer = manufacturer[0]
                device["specific_data.data.network_interfaces.manufacturer"] = manufacturer

            matching_key = next(
                (k for k, v_list in subClassification.items() if manufacturer in v_list),
                "Desconocido"
            )
            device["Clasificacion"] = matching_key

        # --------------------------------------------------
        # EXCEL
        # --------------------------------------------------
        path = (
            f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/'
            f'REPORTE_DISCOVERY_{central}_{current_date_and_time}.xlsx'
        )

        wb = load_workbook(path)
        ws = wb[sheet_name]

        # Limpiar datos anteriores
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.value = None

        headers = ["Adaptadores", "IPs", "MACs", "Manufacturer", "Clasificacion"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=4, column=i, value=h).font = styles.Font(bold=True)

        row_num = 5
        for device in devices:
            for col_num, key in enumerate(columns, 1):
                value = device.get(key, "")
                if isinstance(value, list):
                    value = "\n".join(str(v) for v in value)
                ws.cell(row=row_num, column=col_num, value=value)
            row_num += 1

        copy_style_from_row(ws, source_row=5, start_row=4, end_row=row_num - 1)
        formater(ws)
        wb.save(path)
        wb.close()
