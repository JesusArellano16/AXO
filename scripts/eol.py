import pandas as pd
import shutil
import os
import openpyxl
from openpyxl.styles import Alignment
from copy import copy


def verificar_archivo(central):
    # Ruta de la carpeta donde debe estar
    ruta_base = "AXONIUS_FILES"
    ruta_completa = os.path.join(ruta_base, central)

    # Verificar si la carpeta central existe
    if not os.path.isdir(ruta_completa):
        return False  # La carpeta no existe

    # Verificar si el archivo "vuln.csv" est√° dentro de central
    return f"eol.csv" in os.listdir(ruta_completa)


def Eol(central, current_date_and_time):
    
    if not verificar_archivo(central=central):
        # Crear el archivo .done aunque no exista el eol.csv
        done_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/done/eol_{central}.done'
        os.makedirs(os.path.dirname(done_path), exist_ok=True)  # Crea la carpeta si no existe
        with open(done_path, 'w') as f:
            f.write("done")
        return  # Terminar la funci√≥n sin error
        #exit()
    #print(f'üöÄ Iniciando proceso para EOL en {central}')

    headers = ["Adaptadores", "Preferred Host Name", "Installed Software", "Software Version", "End of Life", "End Of Support", "IPs", "MAC", "Tipo y distribuci√≥n OS", "Cortex", "Virtual Patching"]

    # Copiar el archivo CSV a la carpeta de reportes
    src_path = f'./AXONIUS_FILES/{central}/eol.csv'
    dest_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/eol.csv'
    shutil.copy(src_path, dest_path)

    # Crear archivo Excel
    namew = f'EOL_{central}_{current_date_and_time}'
    name = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/{namew}.xlsx'
    read_file_product = pd.read_csv(dest_path)
    read_file_product.to_excel(name, index=None, header=True)
    os.remove(dest_path)

    # Modificar el archivo Excel
    wb = openpyxl.load_workbook(name)
    ws = wb['Sheet1']
    ws.title = namew
    ws.delete_cols(1)
    for col_num, encabezado in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=encabezado)
    for row in range(2, ws.max_row + 1):
        value_a = str(ws[f"A{row}"].value).lower()  # Leer y convertir a min√∫sculas
        if "paloalto" in value_a:
            ws[f"J{row}"] = "SI"
        elif "paloalto" not in value_a:
            ws[f"J{row}"] = "NO"
        if "deep_security_adapter" in value_a:
            ws[f"K{row}"] = "SI"
        elif "deep_security_adapter" not in value_a:
            ws[f"K{row}"] = "NO"
    if ws["I1"].has_style:
        ws["J1"].font = copy(ws["I1"].font)
        ws["J1"].border = copy(ws["I1"].border)
        ws["J1"].fill = copy(ws["I1"].fill)
        ws["J1"].alignment = copy(ws["I1"].alignment)
        ws["J1"].number_format = copy(ws["I1"].number_format)
        ws["J1"].protection = copy(ws["I1"].protection)
        ws["K1"].font = copy(ws["I1"].font)
        ws["K1"].border = copy(ws["I1"].border)
        ws["K1"].fill = copy(ws["I1"].fill)
        ws["K1"].alignment = copy(ws["I1"].alignment)
        ws["K1"].number_format = copy(ws["I1"].number_format)
        ws["K1"].protection = copy(ws["I1"].protection)


    # Columnas objetivo y anchos
    columnas_objetivo = [1,3,7,8]
    columnas_ancho = {
        "A": 30,
        "B": 45,
        "C": 55,
        "D": 40,
        "E": 20,
        "F": 20,
        "G": 25,
        "H": 20,
        "I": 25,
        "J": 10,
        "K": 20
    }

    # Filtro para las celdas que quieres modificar
    fil = "A1:K1"

    # Iterar sobre las columnas objetivo y aplicar el ajuste de alineaci√≥n "wrap_text"
    for col in columnas_objetivo:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)  # Activar "Wrap Text"

    # Cambiar los anchos de las columnas
    for col, width in columnas_ancho.items():
        ws.column_dimensions[col].width = width

    # Aplicar el filtro a las celdas A1:G1
    ws.auto_filter.ref = fil


    wb.save(name)
    wb.close()
    done_path = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}/done/eol_{central}.done'
    with open(done_path, 'w') as f:
        f.write("done")
    #print(f'‚úÖ Proceso finalizado: {name} creado exitosamente')
