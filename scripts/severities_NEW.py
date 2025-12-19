from axonius_api_client import Connect
import os
from pathlib import Path
from dotenv import load_dotenv
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from collections import defaultdict
import datetime as dt
import warnings
import axonius_api_client


# ======================================================
# WARNINGS OFF
# ======================================================
warnings.filterwarnings(action="ignore", category=axonius_api_client.exceptions.ExtraAttributeWarning)
warnings.filterwarnings("ignore")


# ======================================================
# JSON OFFLINE UTILITIES
# ======================================================
def resolve_json_from_severity(severidad: str):
    sev = severidad.upper()
    if sev == "CRITICAL":
        return "CRITICAL_VULNERABILITIES_GENERAL_SERVERS.json"
    if sev == "HIGH":
        return "HIGH_VULNERABILITIES_GENERAL_SERVERS.json"
    raise ValueError(f"No offline JSON for severity: {severidad}")


def load_vulns_from_json(severidad, central):
    base_dir = Path(__file__).parent.parent / "AXONIUS_FILES" / "GENERAL_JSON"
    json_file = resolve_json_from_severity(severidad)
    json_path = base_dir / json_file

    if not json_path.exists():
        raise FileNotFoundError(json_path)

    with open(json_path, "r", encoding="utf-8") as f:
        devices = json.load(f)

    return [d for d in devices if central in d.get("labels", [])]


# ======================================================
# FUNCIÓN PRINCIPAL
# ======================================================
def get_severities(severidad, central, f_central):
    current_date_and_time = str(dt.date.today())

    from table import mostrar_tabla
    from centrales import centrales
    from critical import critical

    mostrar_tabla(centrales, current_date_and_time)

    base_dir = f'./ARCHIVOS_REPORTES/{central}/{current_date_and_time}'
    os.makedirs(f'{base_dir}/done', exist_ok=True)

    # --------------------------------------------------
    # DATA SOURCE
    # --------------------------------------------------
    dotenv_path = Path(__file__).parent / ".env"
    load_dotenv(dotenv_path=dotenv_path)

    connect_args = {
        "url": os.getenv("AXONIUS_URL"),
        "key": os.getenv("AXONIUS_KEY"),
        "secret": os.getenv("AXONIUS_SECRET"),
        "verify": False
    }

    use_api = (central == "IXTLA")

    if use_api:
        client = Connect(**connect_args)
        apiobj = client.devices
        query_name = f"{severidad} VULNERABILITIES SERVERS IN {f_central}"
        try:
            devices = apiobj.get_by_saved_query(query_name)
        except Exception:
            critical(central=central, current_date_and_time=current_date_and_time, severidad=severidad)
            return
    else:
        if severidad in ("CRITICAL", "HIGH"):
            devices = load_vulns_from_json(severidad, central)
        else:
            client = Connect(**connect_args)
            apiobj = client.devices
            query_name = f"{severidad} VULNERABILITIES SERVERS IN {f_central}"
            devices = apiobj.get_by_saved_query(query_name)

    mostrar_tabla(centrales, current_date_and_time)

    if not devices:
        with open(f'{base_dir}/done/{severidad.lower()}_{central}.done', "w") as f:
            f.write("done")
        return

    # --------------------------------------------------
    # NORMALIZACIÓN (MISMA QUE ANTES)
    # --------------------------------------------------
    filtered_devices = []

    for device in devices:
        filtered_device = {
            "adapters": device.get("adapters", []),
            "hostname": device.get("specific_data.data.hostname_preferred"),
            "ips": device.get("specific_data.data.network_interfaces.ips_preferred", []),
            "macs": device.get("specific_data.data.network_interfaces.mac_preferred", []),
            "os": device.get("specific_data.data.os.type_distribution_preferred"),
            "software_cves": []
        }

        for cve in device.get("specific_data.data.software_cves", []):
            filtered_device["software_cves"].append({
                "cve_id": cve.get("cve_id"),
                "cve_severity": cve.get("cve_severity"),
                "cve_description": cve.get("cve_description")
            })

        filtered_devices.append(filtered_device)

    # --------------------------------------------------
    # EXCEL (FORMATO ORIGINAL + CONTROL DE FILAS)
    # --------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = f"{severidad}_SEV_{central}_{current_date_and_time}"

    MAX_EXCEL_ROWS = 1048576
    current_row = 1
    excel_limit_reached = False

    headers = [
        "Adaptadores", "CVE", "Numero de Dispositivos", "Severidad",
        "Descripcion", "Hostname", "IPs", "MAC",
        "Tipo y distribución OS", "Cortex", "Virtual Patching"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    cve_to_hosts = defaultdict(set)
    cve_desc = {}

    for d in filtered_devices:
        for cve in d["software_cves"]:
            cve_to_hosts[cve["cve_id"]].add(d["hostname"])
            cve_desc[cve["cve_id"]] = cve["cve_description"]

    for d in filtered_devices:
        adapters = "\n".join(d["adapters"])
        ips = "\n".join(d["ips"])
        macs = "\n".join(d["macs"])
        cortex = "SI" if "paloalto_xdr_adapter" in d["adapters"] else "NO"
        vpatch = "SI" if "deep_security_adapter" in d["adapters"] else "NO"

        for cve in d["software_cves"]:
            if current_row >= MAX_EXCEL_ROWS:
                excel_limit_reached = True
                break

            ws.append([
                adapters,
                cve["cve_id"],
                len(cve_to_hosts[cve["cve_id"]]),
                cve["cve_severity"],
                cve["cve_description"],
                d["hostname"],
                ips,
                macs,
                d["os"],
                cortex,
                vpatch
            ])
            current_row += 1

        if excel_limit_reached:
            break

    ws.auto_filter.ref = ws.dimensions

    for col, width in {
        "A": 30, "B": 20, "C": 25, "D": 10, "E": 40,
        "F": 30, "G": 25, "H": 25, "I": 30, "J": 15, "K": 20
    }.items():
        ws.column_dimensions[col].width = width

    for col in ["A", "E", "G", "H"]:
        for cell in ws[col]:
            if cell.row > 1:
                cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

    # --------------------------------------------------
    # HOJA RESUMEN (IGUAL QUE ANTES)
    # --------------------------------------------------
    ws_resumen = wb.create_sheet("RESUMEN")
    ws_resumen.append(["CVE", severidad, "Descripcion"])

    for cve, hosts in cve_to_hosts.items():
        ws_resumen.append([cve, len(hosts), cve_desc.get(cve, "")])

    for cell in ws_resumen[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for col, width in {"A": 30, "B": 20, "C": 40}.items():
        ws_resumen.column_dimensions[col].width = width

    for cell in ws_resumen["C"]:
        if cell.row > 1:
            cell.alignment = Alignment(wrap_text=True)

    # --------------------------------------------------
    # SAVE + DONE
    # --------------------------------------------------
    excel_path = f'{base_dir}/{severidad}_SEV_{central}_{current_date_and_time}.xlsx'
    wb.save(excel_path)

    with open(f'{base_dir}/done/{severidad.lower()}_{central}.done', "w") as f:
        f.write("done")

    mostrar_tabla(centrales, current_date_and_time)


def main():
    get_severities(severidad="CRITICAL", central="POLANCO", f_central="POLANCO")


if __name__ == "__main__":
    main()
