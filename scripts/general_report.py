from pathlib import Path
import shutil
import datetime as dt
from openpyxl import load_workbook
import json
from openpyxl.styles import PatternFill
import urllib3
import warnings
from axonius_api_client import Connect
from dotenv import load_dotenv
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from classifications import subClassification


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.simplefilter("ignore", category=Warning)

def count_unique_servers(results):
    unique_hosts = set()

    for asset in results:
        hostname = asset.get("specific_data.data.hostname_preferred")
        if hostname:
            unique_hosts.add(hostname.strip().upper())

    return len(unique_hosts)

def fetch_and_count(idx, query_name, devices_api):
    results = devices_api.get_by_saved_query(query_name)
    cantidad_unicos = count_unique_servers(results)
    return idx, cantidad_unicos


def run_general_report(central):
    base_dir = Path(__file__).parent.parent
    today = str(dt.date.today())
    current_date_and_time = str(dt.date.today())
    src_file = base_dir / "src" / "Reporte_Discovery.xlsx"

    central_dir = base_dir / "ARCHIVOS_REPORTES" / central
    dated_dir = central_dir / today
    dest_file = dated_dir / f"Reporte_Discovery_{central}_{today}.xlsx"

    # Crear carpeta central si no existe
    central_dir.mkdir(parents=True, exist_ok=True)

    if dated_dir.exists():
        shutil.rmtree(dated_dir)

    dated_dir.mkdir(parents=True, exist_ok=True)

    # Copiar archivo base
    shutil.copy(src_file, dest_file)

    # Abrir Excel
    wb = load_workbook(dest_file)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":  # fórmula
                    cell.value = None       # eliminar fórmula

    ws = wb["Resumen"]

    for row in range(3, 17):
        cell = ws[f"F{row}"]
        cell.value = None
        cell.hyperlink = None

    for row in range(20, 23):
        cell = ws[f"E{row}"]
        cell.value = None
        cell.hyperlink = None

    if central == "GENERAL":
        for sheet_name in wb.sheetnames:
            if sheet_name != "Resumen":
                del wb[sheet_name]
        ws["A2"] = "Inventario - Resumen"
        ws["A19"] = "Servidores - Vulnerabilidades"
    else:
        del wb["Adaptadores integrados"]
        ws["A2"] = f"Inventario {central} - Resumen"
        ws["A19"] = f"Servidores {central} - Vulnerabilidades"

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    max_col = ws.max_column

    for row in (8, 9):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).fill = red_fill

    assets_count = count_assets_by_central(central)
    ws["F3"] = assets_count
    ws["F4"] = "=F5+F10+F13+F14"
    ws["F7"] = "=F5-F6"
    ws["F12"] = "=F10-F11"
    ws["F16"] = "=F3-F4-F15"

    servers_total, servers_xdr = count_assets_and_xdr(
        central,
        "GENERAL_SERVERS.json"
    )

    ws["F5"] = servers_total
    ws["F6"] = servers_xdr
    pcs_total, pcs_xdr = count_assets_and_xdr(
        central,
        "GENERAL_PCs.json"
    )
    ws["F10"] = pcs_total
    ws["F11"] = pcs_xdr
    ws["F13"] = count_assets_simple(
        central,
        "ALL_GENERAL_NETWORK_DEVICES.json"
    )
    ws["F14"] = count_assets_simple(
        central,
        "GENERAL_VARIOUS_IDENTIFIED_DEVICES.json"
    )
    ws["F15"] = count_assets_simple(
        central,
        "ALL_GENERAL_UNIDENTIFIED_SERVERS.json"
    )
    ws["E20"] = count_assets_simple(
        central,
        "CRITICAL_VULNERABILITIES_GENERAL_SERVERS.json"
    )
    ws["E21"] = count_assets_simple(
        central,
        "HIGH_VULNERABILITIES_GENERAL_SERVERS.json"
    )
    ws["E22"] = count_assets_simple(
        central,
        "EoL_GENERAL_SERVERS.json"
    )
    wb.save(dest_file)

    if central == "GENERAL":
        cell_map = {
            0: "E20",
            1: "E21"
        }

        ixtla_file = (
            base_dir
            / "ARCHIVOS_REPORTES"
            / "IXTLA"
            / today
            / f"Reporte_Discovery_IXTLA_{today}.xlsx"
        )
        #"""
        
        if ixtla_file.exists():

            wb_ixtla = load_workbook(ixtla_file, data_only=True)

            ws_inv = wb_ixtla["Inventario"]

            max_row = ws_inv.max_row

            critical = 0
            high = 0

            for row in ws_inv.iter_rows(min_row=6, max_row=max_row, min_col=8, max_col=8):
                value = row[0].value
                if isinstance(value, (int, float)) and value > 0:
                    critical += 1

            for row in ws_inv.iter_rows(min_row=6, max_row=max_row, min_col=9, max_col=9):
                value = row[0].value
                if isinstance(value, (int, float)) and value > 0:
                    high += 1
    

            # Valores actuales del GENERAL
            general_critical = int(ws["E20"].value or 0)
            general_high = int(ws["E21"].value or 0)

            # Sumar
            ws["E20"] = general_critical + critical
            ws["E21"] = general_high + high

            wb_ixtla.close()

        else:
            print(f"⚠️ No existe reporte IXTLA: {ixtla_file}")
        #"""

    else:
        ws = wb["Inventario - PC"]

        base_dir = Path(__file__).parent.parent
        pcs_path = base_dir / "AXONIUS_FILES" / "GENERAL_JSON" / "GENERAL_PCs.json"
        unique_labels_path = base_dir / "scripts" / "UNIQUE_LABELS.json"

        # Obtener número de región (ejemplo R4 → 4)
        try:
            region_number = int(central.replace("R", ""))
        except ValueError:
            return dest_file

        # Cargar labels por región
        with open(unique_labels_path, "r", encoding="utf-8") as f:
            unique_labels = json.load(f)["unique_labels"]

        region_labels = {
            label for label, region in unique_labels.items()
            if region == region_number
        }
        #print(region_labels)
        # Cargar PCs
        with open(pcs_path, "r", encoding="utf-8") as f:
            pcs = json.load(f)

        # Filtrar PCs por región
        filtered_pcs = []

        for asset in pcs:
            asset_labels = asset.get("labels", [])
            asset_id = asset.get("internal_axon_id")

            if not asset_id or not isinstance(asset_labels, list):
                continue

            if central == "R9" and "IXTLAHUACA" in asset_labels:
                continue

            if any(label in region_labels for label in asset_labels):
                filtered_pcs.append(asset)
        #print(len(filtered_pcs))
        # Encabezado
        ws["A3"] = f"PCs {central} - Inventario"

        # Fila base de formato
        base_row = 6

        from openpyxl.styles import Alignment

        for idx, asset in enumerate(filtered_pcs, start=1):
            row = base_row + idx - 1

            # Numeración
            ws[f"A{row}"] = idx

            # Hostname
            hostname = asset.get("specific_data.data.hostname_preferred")
            ws[f"B{row}"] = hostname

            # IPs
            ips = asset.get("specific_data.data.network_interfaces.ips_preferred", [])
            if isinstance(ips, list):
                ips_value = ", ".join(ips)
            else:
                ips_value = ips
            ws[f"C{row}"] = ips_value
            ws[f"C{row}"].alignment = Alignment(wrap_text=True)

            # MACs
            macs = asset.get("specific_data.data.network_interfaces.mac_preferred", [])
            if isinstance(macs, list):
                macs_value = ", ".join(macs)
            else:
                macs_value = macs
            ws[f"D{row}"] = macs_value
            ws[f"D{row}"].alignment = Alignment(wrap_text=True)

            # OS
            os_type = asset.get("specific_data.data.os.type_distribution_preferred")
            ws[f"E{row}"] = os_type

            adapters = asset.get("adapters", [])

            # Palo Alto XDR
            ws[f"F{row}"] = "SI" if "paloalto_xdr_adapter" in adapters else "NO"

            # Deep Security
            ws[f"G{row}"] = "SI" if "deep_security_adapter" in adapters else "NO"

            # Copiar formato de fila 6
            for col in range(1, 8):
                ws.cell(row=row, column=col)._style = ws.cell(row=base_row, column=col)._style
        wb.save(dest_file)
        # =========================
        # INVENTARIO - RED
        # =========================

        ws_red = wb["Inventario - Red"]

        network_path = base_dir / "AXONIUS_FILES" / "GENERAL_JSON" / "ALL_GENERAL_NETWORK_DEVICES.json"

        with open(network_path, "r", encoding="utf-8") as f:
            network_devices = json.load(f)

        filtered_network = []

        for asset in network_devices:
            asset_labels = asset.get("labels", [])
            asset_id = asset.get("internal_axon_id")

            if not asset_id or not isinstance(asset_labels, list):
                continue

            if central == "R9" and "IXTLAHUACA" in asset_labels:
                continue

            if any(label in region_labels for label in asset_labels):
                filtered_network.append(asset)

        # Encabezado
        ws_red["A3"] = f"Network Devices {central} - Inventario"

        from openpyxl.styles import Alignment

        ws_red["A4"] = "Adapters"
        ws_red["A4"].alignment = Alignment(wrap_text=True)
        ws_red.column_dimensions["A"].width = 30

        ws_red["C4"] = "IPs"
        ws_red["E4"] = "OS"
        ws_red["F4"] = "Manufacturer"

        # Fila base de formato (fila 5 del template)
        base_row = 5

        # Limpiar valores existentes desde fila 5
        for row in ws_red.iter_rows(min_row=5, max_row=ws_red.max_row):
            for cell in row:
                cell.value = None

        for idx, asset in enumerate(filtered_network, start=1):
            row = base_row + idx - 1

            # 1️⃣ Copiar formato primero
            for col in range(1, 7):
                ws_red.cell(row=row, column=col)._style = ws_red.cell(row=base_row, column=col)._style

            # 2️⃣ Ahora escribir valores

            # ADAPTERS
            adapters = asset.get("adapters", [])
            if isinstance(adapters, list):
                adapters_unique = list(dict.fromkeys(adapters))
                adapters_value = "\n".join(adapters_unique)
            else:
                adapters_value = adapters

            ws_red[f"A{row}"] = adapters_value
            ws_red[f"A{row}"].alignment = Alignment(wrap_text=True)

            # HOSTNAME
            hostname = asset.get("specific_data.data.hostname_preferred")
            ws_red[f"B{row}"] = hostname

            # IPs
            ips = asset.get("specific_data.data.network_interfaces.ips_preferred", [])
            if isinstance(ips, list):
                ips_value = "\n".join(ips)
            else:
                ips_value = ips

            ws_red[f"C{row}"] = ips_value
            ws_red[f"C{row}"].alignment = Alignment(wrap_text=True)

            # MACs
            macs = asset.get("specific_data.data.network_interfaces.mac_preferred", [])
            if isinstance(macs, list):
                macs_value = "\n".join(macs)
            else:
                macs_value = macs

            ws_red[f"D{row}"] = macs_value
            ws_red[f"D{row}"].alignment = Alignment(wrap_text=True)

            # OS
            os_type = asset.get("specific_data.data.os.type_distribution_preferred")
            ws_red[f"E{row}"] = os_type

            # Manufacturer
            manufacturer = asset.get("specific_data.data.network_interfaces.manufacturer", [])
            if isinstance(manufacturer, list):
                manufacturer_value = "\n".join(manufacturer)
            else:
                manufacturer_value = manufacturer

            ws_red[f"F{row}"] = manufacturer_value
            ws_red[f"F{row}"].alignment = Alignment(wrap_text=True)
                # =========================
        # INVENTARIO - SERVIDORES
        # =========================

        ws_srv = wb["Inventario"]

        servers_path = base_dir / "AXONIUS_FILES" / "GENERAL_JSON" / "GENERAL_SERVERS.json"
        high_path = base_dir / "AXONIUS_FILES" / "GENERAL_JSON" / "HIGH_VULNERABILITIES_GENERAL_SERVERS.json"
        critical_path = base_dir / "AXONIUS_FILES" / "GENERAL_JSON" / "CRITICAL_VULNERABILITIES_GENERAL_SERVERS.json"

        with open(servers_path, "r", encoding="utf-8") as f:
            servers = json.load(f)

        with open(high_path, "r", encoding="utf-8") as f:
            high_vulns = json.load(f)

        with open(critical_path, "r", encoding="utf-8") as f:
            critical_vulns = json.load(f)

        # Filtrar servidores por región
        filtered_servers = []

        for asset in servers:
            asset_labels = asset.get("labels", [])
            asset_id = asset.get("internal_axon_id")

            if not asset_id or not isinstance(asset_labels, list):
                continue

            if central == "R9" and "IXTLAHUACA" in asset_labels:
                continue

            if any(label in region_labels for label in asset_labels):
                filtered_servers.append(asset)

        # Crear mapa hostname -> cantidad HIGH
        high_map = {}
        for asset in high_vulns:
            hostname = asset.get("specific_data.data.hostname_preferred")
            cves = asset.get("specific_data.data.software_cves", [])
            if hostname:
                high_map[hostname.strip().upper()] = len(cves)

        # Crear mapa hostname -> cantidad CRITICAL
        critical_map = {}
        for asset in critical_vulns:
            hostname = asset.get("specific_data.data.hostname_preferred")
            cves = asset.get("specific_data.data.software_cves", [])
            if hostname:
                critical_map[hostname.strip().upper()] = len(cves)

        # Encabezado
        ws_srv["A3"] = f"Servidores {central} - Inventario"

        base_row = 6

        # Limpiar desde fila 6
        for row in ws_srv.iter_rows(min_row=6, max_row=ws_srv.max_row):
            for cell in row:
                cell.value = None

        from openpyxl.styles import Alignment

        for idx, asset in enumerate(filtered_servers, start=1):
            row = base_row + idx - 1

            # 1️⃣ Copiar formato primero
            for col in range(1, 11):
                ws_srv.cell(row=row, column=col)._style = ws_srv.cell(row=base_row, column=col)._style

            # Conteo
            ws_srv[f"A{row}"] = idx

            hostname = asset.get("specific_data.data.hostname_preferred")
            hostname_key = hostname.strip().upper() if hostname else None

            # Hostname
            ws_srv[f"B{row}"] = hostname

            # IPs
            ips = asset.get("specific_data.data.network_interfaces.ips_preferred", [])
            if isinstance(ips, list):
                ips_value = "\n".join(ips)
            else:
                ips_value = ips
            ws_srv[f"C{row}"] = ips_value
            ws_srv[f"C{row}"].alignment = Alignment(wrap_text=True)

            # MACs
            macs = asset.get("specific_data.data.network_interfaces.mac_preferred", [])
            if isinstance(macs, list):
                macs_value = "\n".join(macs)
            else:
                macs_value = macs
            ws_srv[f"D{row}"] = macs_value
            ws_srv[f"D{row}"].alignment = Alignment(wrap_text=True)

            # OS
            os_type = asset.get("specific_data.data.os.type_distribution_preferred")
            ws_srv[f"E{row}"] = os_type

            adapters = asset.get("adapters", [])

            # Palo Alto
            ws_srv[f"F{row}"] = "SI" if "paloalto_xdr_adapter" in adapters else "NO"

            # Deep Security
            ws_srv[f"G{row}"] = "SI" if "deep_security_adapter" in adapters else "NO"

            # HIGH vulnerabilities
            high_count = high_map.get(hostname_key, 0) if hostname_key else 0
            ws_srv[f"H{row}"] = high_count

            # CRITICAL vulnerabilities
            critical_count = critical_map.get(hostname_key, 0) if hostname_key else 0
            ws_srv[f"I{row}"] = critical_count

            # Columna J
            ws_srv[f"J{row}"] = "NA"
                # =========================
        # INVENTARIO - EOL
        # =========================

        ws_eol = wb["Inventario - EOL"]

        eol_path = base_dir / "AXONIUS_FILES" / "GENERAL_JSON" / "EoL_GENERAL_SERVERS.json"

        with open(eol_path, "r", encoding="utf-8") as f:
            eol_servers = json.load(f)

        # Filtrar servidores EOL por región
        filtered_eol = []

        for asset in eol_servers:
            asset_labels = asset.get("labels", [])
            asset_id = asset.get("internal_axon_id")

            if not asset_id or not isinstance(asset_labels, list):
                continue

            if central == "R9" and "IXTLAHUACA" in asset_labels:
                continue

            if any(label in region_labels for label in asset_labels):
                filtered_eol.append(asset)

        # Encabezado
        ws_eol["A3"] = f"Servidores EOL {central} - Inventario"

        base_row = 6

        # Limpiar desde fila 6
        for row in ws_eol.iter_rows(min_row=6, max_row=ws_eol.max_row):
            for cell in row:
                cell.value = None

        from openpyxl.styles import Alignment

        for idx, asset in enumerate(filtered_eol, start=1):
            row = base_row + idx - 1

            # 1️⃣ Copiar formato primero
            for col in range(1, 9):
                ws_eol.cell(row=row, column=col)._style = ws_eol.cell(row=base_row, column=col)._style

            # Conteo
            ws_eol[f"A{row}"] = idx

            # Hostname
            hostname = asset.get("specific_data.data.hostname_preferred")
            ws_eol[f"B{row}"] = hostname

            # IPs
            ips = asset.get("specific_data.data.network_interfaces.ips_preferred", [])
            if isinstance(ips, list):
                ips_value = "\n".join(ips)
            else:
                ips_value = ips

            ws_eol[f"C{row}"] = ips_value
            ws_eol[f"C{row}"].alignment = Alignment(wrap_text=True)

            # MACs
            macs = asset.get("specific_data.data.network_interfaces.mac_preferred", [])
            if isinstance(macs, list):
                macs_value = "\n".join(macs)
            else:
                macs_value = macs

            ws_eol[f"D{row}"] = macs_value
            ws_eol[f"D{row}"].alignment = Alignment(wrap_text=True)

            # OS
            os_type = asset.get("specific_data.data.os.type_distribution_preferred")
            ws_eol[f"E{row}"] = os_type

            adapters = asset.get("adapters", [])

            # Palo Alto
            ws_eol[f"F{row}"] = "SI" if "paloalto_xdr_adapter" in adapters else "NO"

            # Deep Security
            ws_eol[f"G{row}"] = "SI" if "deep_security_adapter" in adapters else "NO"

            # Installed software count
            installed = asset.get("specific_data.data.installed_software", [])
            if isinstance(installed, list):
                ws_eol[f"H{row}"] = len(installed)
            else:
                ws_eol[f"H{row}"] = 0

        #
        # ================================
        # HOJA: Inventario No Identificados
        # ================================
        
        # ================================
        # HOJA: Inventario No Identificados
        # ================================

        ws_unid = wb["Inventario No Identificados"]

        unidentified_path = base_dir / "AXONIUS_FILES" / "GENERAL_JSON" / "ALL_GENERAL_UNIDENTIFIED_SERVERS.json"

        with open(unidentified_path, "r", encoding="utf-8") as f:
            unidentified_servers = json.load(f)

        # Filtrar por región
        filtered_unidentified = []

        for asset in unidentified_servers:
            asset_labels = asset.get("labels", [])
            asset_id = asset.get("internal_axon_id")

            if not asset_id or not isinstance(asset_labels, list):
                continue

            if central == "R9" and "IXTLAHUACA" in asset_labels:
                continue

            if any(label in region_labels for label in asset_labels):
                filtered_unidentified.append(asset)

        # Limpiar desde fila 5
        for row in ws_unid.iter_rows(min_row=5, max_row=ws_unid.max_row):
            for cell in row:
                cell.value = None

        base_row = 5

        from openpyxl.styles import Alignment

        for idx, asset in enumerate(filtered_unidentified, start=1):
            row = base_row + idx - 1

            # Copiar formato base (fila 5 del template)
            for col in range(1, 6):
                ws_unid.cell(row=row, column=col)._style = ws_unid.cell(row=base_row, column=col)._style

            # A -> Adaptadores
            adapters = asset.get("adapters", [])
            if isinstance(adapters, list):
                adapters_value = "\n".join(adapters)
            else:
                adapters_value = adapters

            ws_unid[f"A{row}"] = adapters_value
            ws_unid[f"A{row}"].alignment = Alignment(wrap_text=True)

            # B -> IP
            ips = asset.get("specific_data.data.network_interfaces.ips_preferred", [])
            if isinstance(ips, list):
                ips_value = "\n".join(ips)
            else:
                ips_value = ips

            ws_unid[f"B{row}"] = ips_value
            ws_unid[f"B{row}"].alignment = Alignment(wrap_text=True)

            # C -> MAC
            macs = asset.get("specific_data.data.network_interfaces.mac_preferred", [])
            if isinstance(macs, list):
                macs_value = "\n".join(macs)
            else:
                macs_value = macs

            ws_unid[f"C{row}"] = macs_value
            ws_unid[f"C{row}"].alignment = Alignment(wrap_text=True)

            # D -> Manufacturer
            manufacturers = asset.get("specific_data.data.network_interfaces.manufacturer", [])
            if isinstance(manufacturers, list):
                manufacturer_value = "\n".join(manufacturers)
            else:
                manufacturer_value = manufacturers

            ws_unid[f"D{row}"] = manufacturer_value
            ws_unid[f"D{row}"].alignment = Alignment(wrap_text=True)

            # E -> Clasificación
            classification_value = ""

            if isinstance(manufacturers, list):
                for category, manufacturer_list in subClassification.items():
                    if any(m in manufacturer_list for m in manufacturers):
                        classification_value = category
                        break

            ws_unid[f"E{row}"] = classification_value
            ws_unid[f"E{row}"].alignment = Alignment(wrap_text=True)

        # Ajustar ancho columnas
        ws_unid.column_dimensions["A"].width = 25
        ws_unid.column_dimensions["B"].width = 20
        ws_unid.column_dimensions["C"].width = 25
        ws_unid.column_dimensions["D"].width = 50
        ws_unid.column_dimensions["E"].width = 70

        # ================================
        # HOJA: Inventario Identificados
        # ================================

        ws_ident = wb["Inventario Identificados"]

        identified_path = base_dir / "AXONIUS_FILES" / "GENERAL_JSON" / "GENERAL_VARIOUS_IDENTIFIED_DEVICES.json"

        with open(identified_path, "r", encoding="utf-8") as f:
            identified_devices = json.load(f)

        # Filtrar por región
        filtered_identified = []

        for asset in identified_devices:
            asset_labels = asset.get("labels", [])
            asset_id = asset.get("internal_axon_id")

            if not asset_id or not isinstance(asset_labels, list):
                continue

            if central == "R9" and "IXTLAHUACA" in asset_labels:
                continue

            if any(label in region_labels for label in asset_labels):
                filtered_identified.append(asset)

        # Limpiar desde fila 5
        for row in ws_ident.iter_rows(min_row=5, max_row=ws_ident.max_row):
            for cell in row:
                cell.value = None

        base_row = 5

        from openpyxl.styles import Alignment

        for idx, asset in enumerate(filtered_identified, start=1):
            row = base_row + idx - 1

            # Copiar formato base
            for col in range(1, 6):
                ws_ident.cell(row=row, column=col)._style = ws_ident.cell(row=base_row, column=col)._style

            # A -> Adaptadores
            adapters = asset.get("adapters", [])
            if isinstance(adapters, list):
                adapters_value = "\n".join(adapters)
            else:
                adapters_value = adapters

            ws_ident[f"A{row}"] = adapters_value
            ws_ident[f"A{row}"].alignment = Alignment(wrap_text=True)

            # B -> IP
            ips = asset.get("specific_data.data.network_interfaces.ips_preferred", [])
            if isinstance(ips, list):
                ips_value = "\n".join(ips)
            else:
                ips_value = ips

            ws_ident[f"B{row}"] = ips_value
            ws_ident[f"B{row}"].alignment = Alignment(wrap_text=True)

            # C -> MAC
            macs = asset.get("specific_data.data.network_interfaces.mac_preferred", [])
            if isinstance(macs, list):
                macs_value = "\n".join(macs)
            else:
                macs_value = macs

            ws_ident[f"C{row}"] = macs_value
            ws_ident[f"C{row}"].alignment = Alignment(wrap_text=True)

            # D -> Manufacturer
            manufacturers = asset.get("specific_data.data.network_interfaces.manufacturer", [])
            if isinstance(manufacturers, list):
                manufacturer_value = "\n".join(manufacturers)
            else:
                manufacturer_value = manufacturers

            ws_ident[f"D{row}"] = manufacturer_value
            ws_ident[f"D{row}"].alignment = Alignment(wrap_text=True)

            # E -> Clasificación
            classification_value = ""

            if isinstance(manufacturers, list):
                for category, manufacturer_list in subClassification.items():
                    if any(m in manufacturer_list for m in manufacturers):
                        classification_value = category
                        break

            ws_ident[f"E{row}"] = classification_value
            ws_ident[f"E{row}"].alignment = Alignment(wrap_text=True)

        # Ajustar ancho columnas
        ws_ident.column_dimensions["A"].width = 25
        ws_ident.column_dimensions["B"].width = 20
        ws_ident.column_dimensions["C"].width = 25
        ws_ident.column_dimensions["D"].width = 50
        ws_ident.column_dimensions["E"].width = 70

        wb.save(dest_file)
        

    meses = {
    "01": "Enero",
    "02": "Febrero",
    "03": "Marzo",
    "04": "Abril",
    "05": "Mayo",
    "06": "Junio",
    "07": "Julio",
    "08": "Agosto",
    "09": "Septiembre",
    "10": "Octubre",
    "11": "Noviembre",
    "12": "Diciembre"
    }
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    now = datetime.now()
    year = now.strftime("%Y")
    month_number = now.strftime("%m")
    month_name = meses[month_number]

    dest_dir = os.path.join(
        base_dir,
        "REPORTES_SEMANALES",
        year,
        month_name,
        current_date_and_time  
    )

    os.makedirs(dest_dir, exist_ok=True)
    file_name = os.path.basename(dest_file)
    dest_path = os.path.join(dest_dir, file_name)
    shutil.copy2(dest_file, dest_path)


            

    # Guardar
    wb.save(dest_file)

    return dest_file


def count_assets_by_central(central):
    base_dir = Path(__file__).parent.parent
    general_assets_path = base_dir / "AXONIUS_FILES" / "GENERAL_JSON" / "GENERAL_ASSETS.json"
    unique_labels_path = base_dir / "scripts" / "UNIQUE_LABELS.json"

    with open(general_assets_path, "r", encoding="utf-8") as f:
        assets = json.load(f)

    if central == "GENERAL":
        return len(assets)

    try:
        region_number = int(central.replace("R", ""))
    except ValueError:
        return 0

    with open(unique_labels_path, "r", encoding="utf-8") as f:
        unique_labels = json.load(f)["unique_labels"]

    region_labels = {
        label for label, region in unique_labels.items()
        if region == region_number
    }

    counted_assets = set()

    for asset in assets:
        asset_labels = asset.get("labels", [])
        asset_id = asset.get("internal_axon_id")

        if not asset_id or not isinstance(asset_labels, list):
            continue

        if central == "R9" and "IXTLAHUACA" in asset_labels:
            continue

        # ✅ Contar si tiene AL MENOS un label válido
        if any(label in region_labels for label in asset_labels):
            counted_assets.add(asset_id)

    return len(counted_assets)


def count_assets_and_xdr(central, json_filename):
    base_dir = Path(__file__).parent.parent
    data_path = base_dir / "AXONIUS_FILES" / "GENERAL_JSON" / json_filename
    unique_labels_path = base_dir / "scripts" / "UNIQUE_LABELS.json"

    with open(data_path, "r", encoding="utf-8") as f:
        assets = json.load(f)

    if central == "GENERAL":
        total_assets = set()
        xdr_assets = set()

        for asset in assets:
            asset_id = asset.get("internal_axon_id")
            adapters = asset.get("adapters", [])

            if not asset_id:
                continue

            total_assets.add(asset_id)

            if "paloalto_xdr_adapter" in adapters:
                xdr_assets.add(asset_id)

        return len(total_assets), len(xdr_assets)

    try:
        region_number = int(central.replace("R", ""))
    except ValueError:
        return 0, 0

    with open(unique_labels_path, "r", encoding="utf-8") as f:
        unique_labels = json.load(f)["unique_labels"]

    region_labels = {
        label for label, region in unique_labels.items()
        if region == region_number
    }

    total_assets = set()
    xdr_assets = set()

    for asset in assets:
        asset_id = asset.get("internal_axon_id")
        asset_labels = asset.get("labels", [])
        adapters = asset.get("adapters", [])

        if not asset_id or not isinstance(asset_labels, list):
            continue

        if central == "R9" and "IXTLAHUACA" in asset_labels:
            continue

        if any(label in region_labels for label in asset_labels):
            total_assets.add(asset_id)

            if "paloalto_xdr_adapter" in adapters:
                xdr_assets.add(asset_id)

    return len(total_assets), len(xdr_assets)



def count_assets_simple(central, json_filename):
    base_dir = Path(__file__).parent.parent
    data_path = base_dir / "AXONIUS_FILES" / "GENERAL_JSON" / json_filename
    unique_labels_path = base_dir / "scripts" / "UNIQUE_LABELS.json"

    with open(data_path, "r", encoding="utf-8") as f:
        assets = json.load(f)

    if central == "GENERAL":
        asset_ids = {
            asset.get("internal_axon_id")
            for asset in assets
            if asset.get("internal_axon_id")
        }
        return len(asset_ids)

    try:
        region_number = int(central.replace("R", ""))
    except ValueError:
        return 0

    with open(unique_labels_path, "r", encoding="utf-8") as f:
        unique_labels = json.load(f)["unique_labels"]

    region_labels = {
        label for label, region in unique_labels.items()
        if region == region_number
    }

    counted_assets = set()

    for asset in assets:
        asset_id = asset.get("internal_axon_id")
        asset_labels = asset.get("labels", [])

        if not asset_id or not isinstance(asset_labels, list):
            continue

        if central == "R9" and "IXTLAHUACA" in asset_labels:
            continue

        if any(label in region_labels for label in asset_labels):
            counted_assets.add(asset_id)

    return len(counted_assets)


def main():
    run_general_report("GENERAL")


if __name__ == "__main__":
    main()